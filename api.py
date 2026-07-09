import os
import re
import json
import time
import difflib
from functools import wraps
from typing import Callable, Any, Dict, List, Optional, TypedDict

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field

from llm_interface import get_provider

from motor_determinista import (
    buscar_pdfs,
    indexar_pdfs,
    consultar_excel,
    evaluar_cumplimiento,
)
from conversor_unidades import (
    convertir_unidades,
    convertir_condiciones_referencia,
    _parse_numeric_value,
    _extract_numeric_with_unit,
    _extract_unit_only,
    _normalize_unit,
)
from condiciones_referencia import (
    convertir_a_condiciones_espana,
    _slug_parametro as _slug_param_comb,
    CONDICIONES_PAIS as _COND_PAIS,
)
import fuente_oficial

try:
    from src.llm.prompts import SYSTEM_PROMPT
except Exception:  # fallback si el paquete src no está en el path
    SYSTEM_PROMPT = (
        "Eres el Asistente Experto de Calidad de Gas Natural. Solo tratas la calidad "
        "del gas natural (España, Portugal, Francia, Italia, Alemania, Países Bajos, Bélgica, Noruega, Polonia, Dinamarca, Hungría, Austria, Suiza, Chequia, Grecia, Irlanda, Rumanía, Eslovaquia, Turquía, Reino Unido, UE). Nunca inventas valores "
        "numéricos: los obtienes de las herramientas deterministas. Cita siempre la fuente."
    )

load_dotenv()

# --- Modelo de lenguaje: proveedor abstracto (ver llm_interface.py) ---------
# TODO el acoplamiento a un proveedor concreto (OpenAI, Anthropic, Ollama…) vive
# en `llm_interface.py`. Aquí solo se pide el proveedor configurado y se usa su
# interfaz genérica. Para cambiar de proveedor basta con la variable de entorno
# LLM_PROVIDER (por defecto "openai"); api.py NO necesita cambios.
provider = get_provider()


class PeticionChat(BaseModel):
    """Cuerpo de la petición para `POST /api/chat`."""

    session_id: str = Field(
        ...,
        description=(
            "Identificador de la sesión de conversación. Agrupa los turnos de un mismo "
            "usuario para conservar el historial (memoria de la conversación) y aplicar el "
            "filtro de coherencia entre mensajes. El frontend lo genera una vez y lo reutiliza."
        ),
        examples=["sesion-a1b2c3d4"],
    )
    mensaje: str = Field(
        ...,
        description="Texto de la pregunta del usuario, en lenguaje natural.",
        examples=["¿Cuál es el límite de PCS en Francia?"],
    )

    model_config = {
        "json_schema_extra": {
            "examples": [
                {
                    "session_id": "sesion-a1b2c3d4",
                    "mensaje": "¿Cuál es el límite de PCS en Francia?",
                }
            ]
        }
    }


class RespuestaChat(BaseModel):
    """Respuesta de `POST /api/chat`."""

    respuesta: str = Field(
        ...,
        description="Texto de la respuesta ya redactado, listo para mostrar al usuario.",
        examples=[
            "En Francia el PCS admisible está entre 10,7 y 12,8 kWh/m³ (fuente: GRTgaz)…"
        ],
    )
    modo: str = Field(
        "ia",
        description=(
            "Origen de la respuesta. `ia`: la redactó el modelo de lenguaje. "
            "`determinista`: la resolvió directamente el motor (cifra de la ontología, "
            "fallback sin LLM, o filtro de coherencia entre mensajes)."
        ),
        examples=["ia"],
    )


class StatusResponse(BaseModel):
    """Respuesta de `GET /api/status`: estado operativo del backend."""

    modo: str = Field(
        ...,
        description="`ia` si hay un proveedor LLM operativo; `determinista` si no lo hay.",
        examples=["ia"],
    )
    detalle: str = Field(
        ...,
        description="Descripción legible del estado (proveedor activo o motivo del fallback).",
        examples=["Agente LLM operativo — OpenAI GPT-4o-mini"],
    )


def medir_tiempo(func: Callable[..., Any]) -> Callable[..., Any]:
    @wraps(func)
    async def wrapper(*args: Any, **kwargs: Any) -> Any:
        inicio = time.perf_counter()
        resultado = await func(*args, **kwargs)
        duracion = time.perf_counter() - inicio
        print(f"[medir_tiempo] {func.__name__} tardó {duracion:.3f} segundos")
        return resultado

    return wrapper


def gestionar_errores(func: Callable[..., Any]) -> Callable[..., Any]:
    @wraps(func)
    async def wrapper(*args: Any, **kwargs: Any) -> Any:
        try:
            return await func(*args, **kwargs)
        except Exception as exc:
            print(f"[gestionar_errores] Error interno: {exc}")
            raise HTTPException(status_code=500, detail="Error interno del servidor")

    return wrapper


# Historial de conversación por sesión (lista de mensajes estilo OpenAI).
session_histories: Dict[str, List[Dict[str, Any]]] = {}


class PendingValidation(TypedDict):
    parametro: str
    pais: str
    valor: float


pending_unit_validations: Dict[str, PendingValidation] = {}


def get_session_history(session_id: str) -> List[Dict[str, Any]]:
    if session_id not in session_histories:
        session_histories[session_id] = []
    return session_histories[session_id]


# Tope de mensajes por sesión (~20 turnos): el historial se conserva, pero se acota para
# no crecer sin límite ni desbordar el contexto del LLM.
_MAX_HISTORY_MSGS = 40


def _registrar_turno(session_id: str, mensaje: str, respuesta: str) -> None:
    """Guarda un turno (pregunta + respuesta) en el historial de la sesión.

    Se usa TAMBIÉN para las respuestas deterministas: antes solo se guardaban los turnos
    que pasaban por el LLM, así que el asistente no recordaba las respuestas de cifras.
    """
    history = get_session_history(session_id)
    history.append({"role": "user", "content": mensaje})
    history.append({"role": "assistant", "content": respuesta})
    if len(history) > _MAX_HISTORY_MSGS:
        del history[: len(history) - _MAX_HISTORY_MSGS]


# --- Herramientas deterministas expuestas al LLM (function calling) --------
# Formato "estilo función de OpenAI", usado como formato de intercambio común.
# Cada adaptador de llm_interface.py lo traduce a su proveedor si hace falta.
LLM_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "consultar_excel",
            "description": "Consulta los límites regulatorios de calidad de gas para un parámetro y país.",
            "parameters": {
                "type": "object",
                "properties": {
                    "parametro": {"type": "string", "description": "Parámetro de calidad (p.ej. O2, PCS, Wobbe, S total)."},
                    "pais": {"type": "string", "description": "País/jurisdicción (España, Portugal, Francia, Italia, Alemania, Países Bajos, Bélgica, Noruega, Polonia, Dinamarca, Hungría, Austria, Suiza, Chequia, Grecia, Irlanda, Rumanía, Eslovaquia, Turquía, Reino Unido, UE)."},
                },
                "required": ["parametro", "pais"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "evaluar_cumplimiento",
            "description": (
                "Evalúa si un valor MEDIDO (aportado por el usuario) cumple los límites "
                "regulatorios. ÚSALA SOLO si el usuario da un valor numérico a evaluar. "
                "Si el usuario solo pregunta por el límite/valor de la normativa (sin dar "
                "un valor propio), usa `consultar_excel`; NUNCA inventes un valor."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "parametro": {"type": "string"},
                    "pais": {"type": "string"},
                    "valor": {"type": "number"},
                    "unidad": {"type": "string"},
                },
                "required": ["parametro", "pais", "valor"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "buscar_pdfs",
            "description": "Busca texto relevante dentro de los PDF normativos indexados.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Texto a buscar en los documentos."},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "convertir_unidades",
            "description": (
                "ÚNICA vía autorizada para convertir unidades de forma exacta y determinista. "
                "Úsala SIEMPRE que necesites normalizar energía (MJ/m³ ↔ kWh/m³), temperatura "
                "(K ↔ °C, °F ↔ °C) o concentración (mg/m³ a 15 °C ↔ mg/Nm³ a 0 °C; mg/Nm³ ↔ ppm). "
                "Nunca calcules una conversión por tu cuenta."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "valor": {"type": "number", "description": "Valor numérico a convertir."},
                    "unidad_origen": {"type": "string", "description": "Unidad de partida (p.ej. mg/Nm³, ppm, MJ/m³, K, °F)."},
                    "unidad_destino": {"type": "string", "description": "Unidad de llegada (p.ej. kWh/m³, °C, ppm)."},
                    "parametro": {"type": "string", "description": "Parámetro asociado (PCS, Wobbe, H2S, O2…); se usa para deducir la masa molar en mg/Nm³ ↔ ppm."},
                    "masa_molar": {"type": "number", "description": "Masa molar en g/mol (opcional; solo para mg/Nm³ ↔ ppm si el componente no es conocido)."},
                },
                "required": ["valor", "unidad_origen", "unidad_destino"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "convertir_condiciones_referencia",
            "description": (
                "Lleva un PCS o Índice de Wobbe desde las condiciones de referencia (temperatura "
                "de combustión) de un país a las de España, con los factores de la Tabla A.1 de la "
                "ISO 13443. Portugal usa combustión 25 °C y España 0 °C, por lo que su PCS/Wobbe se "
                "multiplica por 1,0026 para compararlos. Úsala para comparar PCS/Wobbe entre países "
                "con distinta temperatura de combustión. NO la uses para otros parámetros."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "valor": {"type": "number", "description": "Valor a convertir (en kWh/m³; convierte antes la unidad si hace falta)."},
                    "parametro": {"type": "string", "description": "PCS o Wobbe."},
                    "pais_origen": {"type": "string", "description": "País de origen (España, Portugal, Francia, Italia, Alemania, Países Bajos, Bélgica, Noruega, Polonia, Dinamarca, Hungría, Austria, Suiza, Chequia, Grecia, Irlanda, Rumanía, Eslovaquia, Turquía, Reino Unido, UE)."},
                },
                "required": ["valor", "parametro", "pais_origen"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "convertir_condiciones_iso13443",
            "description": (
                "Convierte PCS, PCI, Índice de Wobbe (o volumen/densidad/densidad relativa) entre DOS "
                "condiciones de referencia ARBITRARIAS según la UNE-EN ISO 13443. A diferencia de "
                "convertir_condiciones_referencia (que solo lleva a España), aquí indicas explícitamente "
                "las temperaturas de combustión y medición de origen y destino. Para los pares tabulados "
                "usa los factores literales de la Tabla A.1 (normativos); para el resto, las ecuaciones del "
                "Anexo B. Úsala cuando las condiciones no sean las de un país conocido. No convierte unidades."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "valor": {"type": "number", "description": "Valor a convertir (ya en la unidad correcta)."},
                    "parametro": {"type": "string", "description": "PCS, PCI, Wobbe, volumen, densidad o densidad relativa."},
                    "comb_origen": {"type": "number", "description": "Temperatura de combustión de origen (°C)."},
                    "med_origen": {"type": "number", "description": "Temperatura de medición/volumen de origen (°C)."},
                    "comb_destino": {"type": "number", "description": "Temperatura de combustión de destino (°C)."},
                    "med_destino": {"type": "number", "description": "Temperatura de medición/volumen de destino (°C)."},
                },
                "required": ["valor", "parametro", "comb_origen", "med_origen", "comb_destino", "med_destino"],
            },
        },
    },
]

TOOL_FUNCS: Dict[str, Callable[..., Any]] = {
    # Las consultas normativas usan la FUENTE OFICIAL (ontología) como primaria; el Excel
    # solo como respaldo. Lambdas con enlace tardío (los wrappers se definen más abajo).
    "consultar_excel": lambda **kw: _consultar_norma(kw.get("parametro", ""), kw.get("pais", "")),
    "evaluar_cumplimiento": lambda **kw: _evaluar_norma(kw.get("parametro", ""), kw.get("pais", ""), kw.get("valor"), kw.get("unidad")),
    "convertir_condiciones_referencia": convertir_a_condiciones_espana,
    "convertir_condiciones_iso13443": convertir_condiciones_referencia,
    "buscar_pdfs": buscar_pdfs,
    "convertir_unidades": convertir_unidades,
}


def responder_con_llm(mensaje: str, session_id: str) -> str:
    """Redacta la respuesta con el LLM configurado usando las herramientas deterministas.

    El modelo NUNCA inventa cifras: los números provienen de las herramientas
    (ontología/Excel/PDF). El LLM solo interpreta la pregunta y redacta el
    resultado. El bucle de tool-calling vive en `llm_interface.py`; aquí solo se
    pasa el system prompt, el historial, el mensaje, los esquemas de herramientas
    y el mapa de funciones deterministas.
    """
    history = get_session_history(session_id)
    texto_final = provider.chat(
        system_prompt=SYSTEM_PROMPT,
        history=history,
        user_message=mensaje,
        tools=LLM_TOOLS,
        tool_functions=TOOL_FUNCS,
        temperature=0,
        max_tool_iterations=5,
    )
    # Persistir el turno en el historial de la sesión.
    _registrar_turno(session_id, mensaje, texto_final)
    return texto_final


backend_mode = "ia" if provider.is_available() else "determinista"
backend_detail = (
    f"Agente LLM operativo — {provider.display_name()}" if backend_mode == "ia"
    else "Sin LLM disponible: usando fallback determinista"
)

API_VERSION = "0.1.0"

API_DESCRIPTION = """
API del **Comparador de Calidad de Gas Natural en Europa**.

Compara los límites regulatorios de calidad del gas natural entre **España** (base de
referencia) y 20 países/regiones europeas, sobre **10 parámetros**: índice de Wobbe, PCS,
densidad relativa, azufre total, H₂S+COS, mercaptanos (RSH), O₂, CO₂ y puntos de rocío de
agua e hidrocarburos.

### Principio de diseño: cero cifras inventadas
* **Motor determinista** (código + ontología validada a partir de PDFs oficiales): única
  fuente autorizada de cifras, límites, conversiones de unidad y normalización de
  condiciones (ISO 13443). Nunca improvisa un número.
* **Capa conversacional** (LLM): interpreta la pregunta y **redacta** la respuesta, pero
  tiene prohibido generar cifras — las obtiene llamando a las herramientas deterministas.

### Modos de respuesta
El backend opera con LLM (`modo: "ia"`) o, si no hay proveedor disponible o el LLM falla,
cae automáticamente al motor determinista (`modo: "determinista"`) sin romper el chat.

### Recorrido de una consulta
`index.html` → `POST /api/chat` → router determinista → (ontología · RAG · LLM) → respuesta citada.
"""

TAGS_METADATA = [
    {"name": "Interfaz", "description": "Sirve la interfaz web del chatbot (HTML/JS)."},
    {"name": "Estado", "description": "Salud y modo operativo del backend (con o sin LLM)."},
    {"name": "Chat", "description": "Conversación en lenguaje natural con el asistente experto."},
    {
        "name": "Comparativa",
        "description": (
            "Datos estructurados para la comparación regulatoria: catálogo de parámetros y "
            "países, comparación puntual y matriz comparativa (heatmap)."
        ),
    },
]

app = FastAPI(
    title="Comparador de Calidad de Gas Natural — API",
    summary="Comparativa regulatoria de calidad del gas natural en Europa (España base · 20 países · 10 parámetros).",
    description=API_DESCRIPTION,
    version=API_VERSION,
    contact={"name": "Equipo ENAG-S_G1 (Enagás)", "email": "ciclabenagas@gmail.com"},
    openapi_tags=TAGS_METADATA,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_INDEX_HTML = os.path.join(os.path.dirname(__file__), "index.html")


@app.get(
    "/",
    tags=["Interfaz"],
    summary="Interfaz web del chatbot",
    response_class=FileResponse,
    response_description="Documento HTML de la interfaz (`index.html`).",
    responses={200: {"content": {"text/html": {}}, "description": "Página del chatbot."}},
)
async def servir_chat() -> FileResponse:
    """Sirve la interfaz web del chatbot en la raíz (`http://localhost:8000/`).

    Devuelve `index.html` con cabeceras **sin caché**: así los compañeros ven SIEMPRE la
    última versión tras un `git pull`, sin tener que forzar recarga (Ctrl+F5) en el navegador.
    """
    return FileResponse(
        _INDEX_HTML,
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


def _normalize_country(text: str) -> Optional[str]:
    normalized = text.lower()
    aliases = {
        "espa": "España",
        "espana": "España",
        "españa": "España",
        "portugal": "Portugal",
        "francia": "Francia",
    }
    for key, value in aliases.items():
        if key in normalized:
            return value
    return None


# Dígitos en subíndice (H₂S, CO₂, O₂) y superíndice (Nm³) → dígitos ASCII.
_SUB_SUP_DIGITS = str.maketrans({
    "₀": "0", "₁": "1", "₂": "2", "₃": "3", "₄": "4",
    "₅": "5", "₆": "6", "₇": "7", "₈": "8", "₉": "9",
    "⁰": "0", "¹": "1", "²": "2", "³": "3", "⁴": "4",
    "⁵": "5", "⁶": "6", "⁷": "7", "⁸": "8", "⁹": "9",
})


def _normalize_parameter(text: str) -> Optional[str]:
    normalized = text.lower()
    # H₂S, CO₂, O₂… escritos con dígitos en subíndice: pásalos a "h2s", "co2", "o2".
    normalized = normalized.translate(_SUB_SUP_DIGITS)
    # "02" (cero-dos) escrito como O2 (oxígeno): normalizar el token aislado.
    # No toca "1.02", "2002" ni "co2" (la barrera \w/.,/ lo evita).
    normalized = re.sub(r"(?<![\w./,])02(?![\w])", "o2", normalized)
    # Punto de rocío: desambiguar HC (hidrocarburos) frente a H2O (agua).
    # Sin esto, "rocío de HC" caía siempre en H2O por el alias genérico "rocío".
    na = normalized.translate(str.maketrans("áéíóúü", "aeiouu"))
    if "rocio" in na or "dew point" in na:
        if "hc" in na or "hidrocarbur" in na:
            return "hc(rocío)"
        return "h2o(rocío)"
    aliases = {
        "o2": "o2",
        "oxigeno": "o2",
        "oxígeno": "o2",
        "pcs": "pcs",
        "h2s": "h2s+cos",
        "h2s+cos": "h2s+cos",
        "wobbe": "wobbe",
        "s total": "s total",
        "co2": "co2",
        "h2o": "h2o(rocío)",
        "h2o(rocío)": "h2o(rocío)",
        "h2o(rocio)": "h2o(rocío)",
        "rocio": "h2o(rocío)",
        "rocío": "h2o(rocío)",
        "hc": "hc(rocío)",
        "hc(rocío)": "hc(rocío)",
        "rsh": "rsh",
        "densidad relativa": "densidad relativa",
        "hco": "hco",
        "indice de wobbe": "wobbe",
        "índice de wobbe": "wobbe",
        "azufre total": "s total",
        "azufre": "s total",
    }
    # Del alias más largo al más corto: evita que un alias corto contenido en otro
    # más largo gane por error (p. ej. "o2" dentro de "co2", o "hc" dentro de "hco").
    for key in sorted(aliases, key=len, reverse=True):
        if key in normalized:
            return aliases[key]
    # Fallback: "S" suelto = símbolo del azufre total (no la "s" interna de otras palabras).
    if re.search(r"(?<![\w])s(?![\w])", normalized):
        return "s total"
    return None


# Slugs exactos del BIOMETANO (los que envía el desplegable). Se comprueban primero,
# para que "co" no choque con "co2" por subcadena.
_SLUGS_BIOMETANO = {"ch4_min", "o2", "co2", "co", "s total", "h2s+cos", "h2o(rocío)",
                    "siloxanos", "comp_oil", "aminas", "nh3", "halogenados"}
# Alias específicos del BIOMETANO (parámetros que no existen en gas natural). Mapa
# propio para no contaminar `_normalize_parameter` del gas natural (riesgo R2).
_ALIASES_BIOMETANO = {
    "ch4": "ch4_min", "ch4_min": "ch4_min", "metano": "ch4_min", "methane": "ch4_min", "methan": "ch4_min",
    "siloxano": "siloxanos", "siloxanos": "siloxanos", "siloxane": "siloxanos",
    "silicio": "siloxanos", "silicium": "siloxanos", "silicon": "siloxanos", "silizium": "siloxanos",
    "aceite": "comp_oil", "compresor": "comp_oil", "comp_oil": "comp_oil", "compressor": "comp_oil",
    "amina": "aminas", "aminas": "aminas", "amine": "aminas", "ammine": "aminas",
    "monoxido de carbono": "co", "carbon monoxide": "co", "monoxide de carbone": "co", "kohlenmonoxid": "co",
    "nh3": "nh3", "amoniaco": "nh3", "amonio": "nh3", "ammonia": "nh3", "ammoniak": "nh3",
    "halogen": "halogenados", "halogenados": "halogenados", "halogenado": "halogenados",
    "cloro": "halogenados", "fluor": "halogenados", "chlorine": "halogenados", "fluorine": "halogenados",
}


def _normalize_parameter_biometano(text: str) -> Optional[str]:
    """Normaliza el nombre de un componente a un slug de BIOMETANO. Primero el slug
    exacto del desplegable, luego los alias específicos del biometano, y por último
    los parámetros que SOLAPAN con el gas natural (O₂, CO₂, azufre, H₂S+COS, rocío)."""
    raw = (text or "").strip().lower()
    if raw in _SLUGS_BIOMETANO:
        return raw
    na = text.lower().translate(_SUB_SUP_DIGITS).translate(str.maketrans("áéíóúü", "aeiouu"))
    for key in sorted(_ALIASES_BIOMETANO, key=len, reverse=True):
        if key in na:
            return _ALIASES_BIOMETANO[key]
    slug = _normalize_parameter(text)
    if slug in {"o2", "co2", "s total", "h2s+cos", "h2o(rocío)"}:
        return slug
    return None


# Slugs y alias del HIDRÓGENO (ISO 14687 Grade D). El desplegable envía el slug exacto.
_SLUGS_HIDROGENO = {"h2_pureza", "no_h2", "h2o", "thc", "ch4", "o2", "he", "n2", "ar", "co2",
                    "co", "s total", "hcho", "hcooh", "nh3", "halogenados", "particulas"}
_ALIASES_HIDROGENO = {
    "pureza": "h2_pureza", "purity": "h2_pureza", "reinheit": "h2_pureza", "purete": "h2_pureza",
    "gases no hidrogeno": "no_h2", "non-hydrogen": "no_h2",
    "agua": "h2o", "water": "h2o",
    "hidrocarburos": "thc", "hydrocarbons": "thc",
    "metano": "ch4", "methane": "ch4",
    "oxigeno": "o2", "oxygen": "o2", "sauerstoff": "o2",
    "helio": "he", "helium": "he",
    "nitrogeno": "n2", "nitrogen": "n2",
    "argon": "ar",
    "dioxido de carbono": "co2", "carbon dioxide": "co2",
    "monoxido de carbono": "co", "carbon monoxide": "co",
    "azufre": "s total", "sulfur": "s total", "sulphur": "s total",
    "formaldehido": "hcho", "formaldehyde": "hcho",
    "acido formico": "hcooh", "formic acid": "hcooh",
    "amoniaco": "nh3", "ammonia": "nh3",
    "halogen": "halogenados", "halide": "halogenados", "cloruro": "halogenados",
    "particulas": "particulas", "particulates": "particulas", "polvo": "particulas",
}


def _normalize_parameter_hidrogeno(text: str) -> Optional[str]:
    """Normaliza el nombre de un componente a un slug de HIDRÓGENO (ISO 14687 Grade D)."""
    raw = (text or "").strip().lower()
    if raw in _SLUGS_HIDROGENO:
        return raw
    na = text.lower().translate(_SUB_SUP_DIGITS).translate(str.maketrans("áéíóúü", "aeiouu"))
    for key in sorted(_ALIASES_HIDROGENO, key=len, reverse=True):
        if key in na:
            return _ALIASES_HIDROGENO[key]
    return None


def _normalize_condition_text(text: Optional[str]) -> str:
    if not text:
        return ""
    cleaned = str(text).strip()
    cleaned = cleaned.replace("Condiciones de medición:", "")
    cleaned = cleaned.replace("Condiciones de medicion:", "")
    cleaned = cleaned.replace("Condiciones:", "")
    return cleaned.strip()


# --- Strict measurement-unit validation dictionary ---
VALIDATION_UNITS = {
    "Índice de Wobbe": "kWh/m³",
    "PCS": "kWh/m³",
    "S": "mg/m³",
    "H2S + COS + RSH": "mg/m³",
    "O2": "% molar",
    "CO2": "% molar",
    "Temperatura de rocío del H2O": "°C",
    "Temperatura de rocío de HC": "°C",
}

DISPLAY_MAP = {
    "wobbe": "Índice de Wobbe",
    "pcs": "PCS",
    "s total": "S",
    "h2s+cos": "H2S + COS + RSH",
    "o2": "O2",
    "co2": "CO2",
    "h2o(rocío)": "Temperatura de rocío del H2O",
    "hc(rocío)": "Temperatura de rocío de HC",
}


def _unit_matches_expected(param: str, unit: Optional[str]) -> bool:
    if not param or not unit:
        return False
    expected = VALIDATION_UNITS.get(DISPLAY_MAP.get(param, param))
    if expected is None:
        return False
    # Acepta la unidad si es la esperada O si es convertible de forma determinista a ella
    # (p.ej. MJ/m³ para Wobbe, ppm para H2S, mg/Nm³ para S…). El conversor decide.
    if _normalize_unit(unit) == _normalize_unit(expected):
        return True
    conv = convertir_unidades(1.0, unit, expected, param)
    return "valor_convertido" in conv


def _expected_unit_for_parameter(param: str) -> str:
    return VALIDATION_UNITS.get(DISPLAY_MAP.get(param, param), "")


def _missing_unit_message(parametro: str) -> str:
    param_display = DISPLAY_MAP.get(parametro, parametro)
    return f"⚠️ Valor detectado sin unidades. Por favor, indícame en qué unidades estás expresando este valor para el parámetro {param_display}."


def _incorrect_unit_message(parametro: str) -> str:
    param_display = DISPLAY_MAP.get(parametro, parametro)
    expected_unit = _expected_unit_for_parameter(parametro)
    return f"❌ Unidades incorrectas. Para el parámetro {param_display}, la unidad requerida es {expected_unit}."


# Lista de parámetros que el sistema sabe consultar (ámbito del PROMPT MAESTRO).
# Se muestra al usuario cuando escribe un índice que no reconocemos.
PARAMETROS_DISPONIBLES = [
    "Índice de Wobbe",
    "PCS (Poder Calorífico Superior)",
    "Densidad relativa",
    "Azufre total (S)",
    "H₂S + COS",
    "Mercaptanos (RSH)",
    "O₂ (oxígeno)",
    "CO₂",
    "Punto de rocío del agua (H₂O)",
    "Punto de rocío de hidrocarburos (HC)",
]


def _parametro_no_reconocido_message() -> str:
    opciones = "\n".join(f"- {p}" for p in PARAMETROS_DISPONIBLES)
    return (
        "No he reconocido el parámetro de tu consulta. "
        "Los parámetros disponibles son:\n\n"
        f"{opciones}\n\n"
        "Indícame uno de ellos junto con el país (España, Portugal, Francia, Italia, Alemania, Países Bajos, Bélgica, Noruega, Polonia, Dinamarca, Hungría, Austria, Suiza, Chequia, Grecia, Irlanda, Rumanía, Eslovaquia, Turquía, Reino Unido o UE) "
        "para darte los valores o comprobar el cumplimiento."
    )


def _mensaje_capacidades() -> str:
    opciones = "\n".join(f"- {p}" for p in PARAMETROS_DISPONIBLES)
    return (
        "Puedo ayudarte con consultas sobre **calidad del gas natural** en España, "
        "Portugal, Francia, Italia, Alemania, Países Bajos, Bélgica, Noruega, Polonia, Dinamarca, Hungría, Austria, Suiza, Chequia, Grecia, Irlanda, Rumanía, Eslovaquia, Turquía, Reino Unido y la UE. Los parámetros que puedes consultar son:\n\n"
        f"{opciones}\n\n"
        "Puedes pedirme, por ejemplo: los valores de un parámetro en un país, "
        "comprobar si un valor cumple la normativa, o comparar dos países."
    )


def _mensaje_fuera_de_ambito() -> str:
    return (
        "Este chat no admite respuestas para ese tipo de preguntas. "
        "Solo respondo a consultas sobre **calidad del gas natural**: introduce un "
        "índice o parámetro de calidad del gas (Índice de Wobbe, PCS, O₂, CO₂, azufre, "
        "punto de rocío…) y, si quieres, un país y un valor."
    )


def _es_pregunta_capacidades(texto_norm: str) -> bool:
    """¿El usuario pregunta qué puede hacer/consultar el chatbot?"""
    patrones = (
        "que puedo consultar", "qué puedo consultar", "que puedo preguntar",
        "qué puedo preguntar", "que se puede consultar", "qué se puede consultar",
        "que valores se pueden consultar", "qué valores se pueden consultar",
        "que valores puedo", "qué valores puedo", "que parametros", "qué parámetros",
        "que parámetros", "qué parametros", "que indices", "qué índices",
        "que índices", "qué indices", "para que sirve", "para qué sirve",
        "que haces", "qué haces", "que puedes hacer", "qué puedes hacer",
        "como funciona", "cómo funciona", "que datos", "qué datos",
        "que preguntas puedo", "qué preguntas puedo", "opciones disponibles",
    )
    return any(p in texto_norm for p in patrones)


def _es_tema_calidad_gas(texto_norm: str) -> bool:
    """¿El mensaje trata, aunque sea vagamente, de calidad del gas natural?"""
    terminos = (
        "gas", "calidad", "wobbe", "pcs", "poder calorifico", "poder calorífico",
        "azufre", "sulfur", "h2s", "cos", "mercaptano", "rsh", "oxigeno", "oxígeno",
        "o2", "co2", "dioxido", "dióxido", "carbono", "rocio", "rocío", "densidad",
        "indice", "índice", "ppm", "nm3", "nm³", "kwh", "mj", "molar",
        "limite", "límite", "normativa", "especificac", "hidrocarburo",
    )
    return any(t in texto_norm for t in terminos)


def _evaluate_validated_comparison(parametro: str, pais: str, valor: float, unidad: str) -> str:
    # Filtrado estricto: solo el país pedido (España se usa por detrás para comparar).
    return _evaluar_paises(parametro, valor, unidad, [pais])


ALL_COUNTRIES = ["España", "Portugal", "Francia", "Italia", "Alemania", "Países Bajos", "Bélgica", "Noruega", "Polonia", "Dinamarca", "Hungría", "Austria", "Suiza", "Chequia", "Grecia", "Irlanda", "Rumanía", "Eslovaquia", "Turquía", "Reino Unido", "UE"]
PAIS_BASE = "España"


def _num_simple(x: Any) -> Optional[float]:
    s = str(x).replace(",", ".")
    m = re.search(r"[-+]?[0-9]*\.?[0-9]+", s)
    return float(m.group(0)) if m else None


def _detectar_discrepancia(rec_oficial: Dict[str, Any], excel: Optional[Dict[str, Any]]) -> str:
    """Compara el límite OFICIAL con el del Excel (mismo parámetro/país). Nota si difieren
    (solo si las unidades coinciden y ambos son numéricos)."""
    if not excel or not excel.get("matches"):
        return ""
    m = excel["matches"][0]
    u_excel = _normalize_unit(_txt(m.get("unidad")).strip("()"))
    u_ofi = _normalize_unit(rec_oficial.get("unidad") or "")
    if u_excel and u_ofi and u_excel != u_ofi:
        return ""  # unidades distintas: no comparamos crudo
    difs = []
    for campo, etiqueta in (("limite_superior", "máximo"), ("limite_inferior", "mínimo")):
        o = _num_simple(rec_oficial.get(campo)); e = _num_simple(_txt(m.get(campo)))
        if o is not None and e is not None and abs(o - e) > 1e-6:
            difs.append(f"{etiqueta}: oficial {rec_oficial.get(campo)} vs Excel {_txt(m.get(campo))}")
    return "; ".join(difs)


def _consultar_norma(parametro: str, pais: str, tipo_gas: str = "gas_natural") -> Dict[str, Any]:
    """Consulta normativa. FUENTE PRIMARIA: documentación oficial (ontología verificada
    de los PDFs en data/raw). El Excel solo como índice/respaldo. Si hay dato oficial y
    el Excel discrepa, prevalece el oficial y se marca la discrepancia.

    `tipo_gas` por defecto "gas_natural" → comportamiento idéntico al actual. Para
    "biometano" se consulta `parametros_biometano` y se OMITE el respaldo del Excel
    (el Excel es solo de gas natural — riesgo R3)."""
    oficial = fuente_oficial.consultar(parametro, pais, tipo_gas)
    if tipo_gas != "gas_natural":
        return oficial  # sin respaldo Excel para gases no naturales
    try:
        excel = consultar_excel(parametro, pais)
    except Exception:
        excel = {"count": 0, "matches": []}
    if oficial.get("count"):
        disc = _detectar_discrepancia(oficial["matches"][0], excel)
        if disc:
            oficial["matches"][0]["discrepancia"] = disc
        return oficial
    return excel  # el oficial no cubre este caso → respaldo del Excel


def _evaluar_norma(parametro: str, pais: str, valor: float, unidad: Optional[str] = None) -> Dict[str, Any]:
    """Evaluación de cumplimiento contra el límite OFICIAL; Excel solo como respaldo."""
    oficial = fuente_oficial.evaluar(parametro, pais, valor, unidad)
    if oficial.get("count"):
        try:
            disc = _detectar_discrepancia(oficial["matches"][0], consultar_excel(parametro, pais))
            if disc:
                oficial["matches"][0]["discrepancia"] = disc
        except Exception:
            pass
        return oficial
    return evaluar_cumplimiento(parametro, pais, valor, unidad=unidad)


def _cita_oficial(item: Dict[str, Any]) -> str:
    """Cita completa de un registro: norma · organismo · fecha · artículo/página — URL."""
    doc = item.get("documento") or "Fuente no especificada"
    partes = [f"**{doc}**"]
    for campo in ("organismo", "fecha", "articulo"):
        v = (item.get(campo) or "").strip()
        if v:
            partes.append(v)
    linea = " · ".join(partes)
    url = (item.get("url") or item.get("pdf") or "").strip()
    if url:
        linea += f" — {url}"
    estado = item.get("estado") or ""
    if estado and estado != "VERIFICADO":
        linea += f" _({estado})_"
    return linea


def _norm_pais(p: Any) -> str:
    s = str(p).strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")):
        s = s.replace(a, b)
    return s


def _txt(v: Any) -> str:
    """Coerciona a texto de forma robusta (NaN de pandas / None -> '')."""
    if v is None:
        return ""
    if isinstance(v, float) and v != v:  # NaN
        return ""
    return str(v).strip()


def _sin_limite(s: str) -> bool:
    """¿La celda indica ausencia de límite numérico?"""
    s = s.lower()
    return (s in ("-", "") or "especific" in s or s.startswith("sin")
            or "no regulad" in s or "monitor" in s or "incluido" in s or "no es fijo" in s)


def _unidad_de_pais(parametro: str, pais: str) -> Optional[str]:
    """Devuelve la unidad que exige la normativa de `pais` para `parametro`."""
    try:
        resp = _consultar_norma(parametro, pais)
    except Exception:
        return None
    for m in resp.get("matches", []):
        u = (m.get("unidad") or m.get("unidad_registro") or "").strip()
        if u:
            return u
    return None


def _estado_comparabilidad(parametro: str, unidad_es: Optional[str], unidad_pais: Optional[str]) -> str:
    """Compara la normativa española vs la del país por su unidad/magnitud.

    - 'Directamente Comparable': misma unidad (o equivalente, factor 1).
    - 'Comparable con Normalización': unidades distintas, misma magnitud física.
    - 'No Comparable': magnitudes incompatibles o falta de datos.
    """
    if not unidad_es or not unidad_pais:
        return "No Comparable"
    conv = convertir_unidades(1.0, unidad_pais, unidad_es, parametro)
    if "valor_convertido" not in conv:
        return "No Comparable"
    if "Sin conversión" in conv.get("formula", ""):
        return "Directamente Comparable"
    return "Comparable con Normalización"


def _celda_es_vs_pais(parametro: str, pais: str, unidad_pais: Optional[str], unidad_es: Optional[str]) -> str:
    """Texto de la celda de comparabilidad cruzada: 'España vs [País]: [Estado]'."""
    if _norm_pais(pais) == _norm_pais(PAIS_BASE):
        return "— (base de referencia)"
    estado = _estado_comparabilidad(parametro, unidad_es, unidad_pais)
    return f"España vs {pais}: {estado}"


_PAIS_FUZZY = {"espana": "España", "portugal": "Portugal", "francia": "Francia",
               "italia": "Italia", "alemania": "Alemania",
               "holanda": "Países Bajos", "nederland": "Países Bajos",
               "belgica": "Bélgica", "belgium": "Bélgica",
               "noruega": "Noruega", "norway": "Noruega",
               "polonia": "Polonia", "poland": "Polonia",
               "dinamarca": "Dinamarca", "denmark": "Dinamarca",
               "hungria": "Hungría", "hungary": "Hungría",
               "austria": "Austria", "suiza": "Suiza", "switzerland": "Suiza",
               "chequia": "Chequia", "czech": "Chequia", "grecia": "Grecia", "greece": "Grecia",
               "irlanda": "Irlanda", "ireland": "Irlanda", "rumania": "Rumanía", "romania": "Rumanía",
               "eslovaquia": "Eslovaquia", "slovakia": "Eslovaquia", "turquia": "Turquía", "turkey": "Turquía",
               "reino unido": "Reino Unido", "united kingdom": "Reino Unido"}


def _detectar_paises(texto_norm: str) -> list:
    """Devuelve la lista de países mencionados, tolerando erratas (ej. 'frnacia').

    Vacía si el usuario no menciona ningún país (→ se asumirán todos).
    """
    t = _norm_pais(texto_norm)  # minúsculas sin acentos
    encontrados: list = []
    # 1) coincidencia directa por subcadena (evita falsos positivos con "espan")
    for kw, nombre in [("espan", "España"), ("spain", "España"),
                       ("portugal", "Portugal"), ("francia", "Francia"), ("france", "Francia"),
                       ("italia", "Italia"), ("italy", "Italia"),
                       ("alemania", "Alemania"), ("germania", "Alemania"), ("germany", "Alemania"), ("deutschland", "Alemania"),
                       ("paises bajos", "Países Bajos"), ("holanda", "Países Bajos"), ("netherlands", "Países Bajos"), ("nederland", "Países Bajos"),
                       ("belgica", "Bélgica"), ("belgium", "Bélgica"), ("belgique", "Bélgica"), ("belgie", "Bélgica"),
                       ("noruega", "Noruega"), ("norway", "Noruega"), ("norge", "Noruega"),
                       ("polonia", "Polonia"), ("poland", "Polonia"), ("polska", "Polonia"),
                       ("dinamarca", "Dinamarca"), ("denmark", "Dinamarca"), ("danmark", "Dinamarca"),
                       ("hungria", "Hungría"), ("hungary", "Hungría"), ("magyar", "Hungría"),
                       ("austria", "Austria"), ("osterreich", "Austria"),
                       ("suiza", "Suiza"), ("switzerland", "Suiza"), ("schweiz", "Suiza"), ("suisse", "Suiza"),
                       ("chequia", "Chequia"), ("republica checa", "Chequia"), ("checa", "Chequia"), ("czech", "Chequia"),
                       ("grecia", "Grecia"), ("greece", "Grecia"),
                       ("irlanda", "Irlanda"), ("ireland", "Irlanda"),
                       ("rumania", "Rumanía"), ("romania", "Rumanía"),
                       ("eslovaquia", "Eslovaquia"), ("slovakia", "Eslovaquia"),
                       ("turquia", "Turquía"), ("turkey", "Turquía"), ("turkiye", "Turquía"),
                       ("reino unido", "Reino Unido"), ("united kingdom", "Reino Unido"), ("gran bretana", "Reino Unido"),
                       ("europa", "UE"), ("union europea", "UE"), ("easee", "UE")]:
        if kw in t and nombre not in encontrados:
            encontrados.append(nombre)
    # "UE" suelto (no la "ue" interna de "fuente", "que", "pequeño"…)
    if "UE" not in encontrados and re.search(r"(?<![\w])ue(?![\w])", t):
        encontrados.append("UE")
    if encontrados:
        return encontrados
    # 2) coincidencia difusa por palabra (tolera erratas: frnacia, portgal, espanha…)
    for palabra in re.findall(r"[a-z]{4,}", t):
        match = difflib.get_close_matches(palabra, list(_PAIS_FUZZY.keys()), n=1, cutoff=0.78)
        if match:
            nombre = _PAIS_FUZZY[match[0]]
            if nombre not in encontrados:
                encontrados.append(nombre)
    return encontrados


def _evaluar_paises(parametro: str, valor: float, unidad: Optional[str], paises: list, todos: bool = False) -> str:
    """Evalúa un valor contra los países indicados y devuelve la tabla.

    FILTRADO ESTRICTO: solo se muestran filas de los países pedidos. España se
    consulta SIEMPRE por detrás (en memoria) para la columna 'Comparabilidad
    normativa', pero NO aparece como fila salvo que el usuario la pida.
    """
    unidad_es = _unidad_de_pais(parametro, PAIS_BASE)  # background: solo para comparar
    filas: list = []
    for pais in paises:
        resp = _evaluar_norma(parametro, pais, valor, unidad=unidad)
        if resp.get("error"):
            continue
        filas.extend(resp.get("matches", []))
    if not filas:
        return (
            f"No encontré registros de '{parametro}' en {', '.join(paises)} "
            f"para evaluar el valor {valor}{f' {unidad}' if unidad else ''}."
        )

    if todos:
        titulo = f"**¿En qué países cumple {parametro} = {valor}{f' {unidad}' if unidad else ''}?**"
    else:
        titulo = f"**Evaluación de cumplimiento — {', '.join(paises)}**"
    # UNA SOLA TABLA: cumplimiento + límite + condiciones de referencia + comparabilidad
    # en la misma fila (las condiciones van AL LADO de los valores, no debajo).
    lines = [
        titulo,
        "",
        "| País | Parámetro | Valor evaluado | Límite normativo | Condiciones de referencia | Resultado | Detalle | Comparabilidad |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    conversiones: list = []
    evidencias: list = []
    cumple_en: list = []
    for item in filas[:12]:
        pais_fila = item.get("pais", "")
        nombre = str(item.get("parametro") or parametro).strip()
        estado = item.get("cumple", "No evaluable")
        detalle = item.get("detalle", "")
        inf = _txt(item.get("limite_inferior")) or "-"
        sup = _txt(item.get("limite_superior")) or "-"
        unidad_reg = item.get("unidad_registro") or unidad or ""
        condiciones = _normalize_condition_text(item.get("condiciones") or item.get("condiciones de medicion") or item.get("condiciones de medición"))
        condiciones = condiciones.replace("bars", "bar").strip() if condiciones else "—"
        res = "🟢 Cumple" if estado == "Cumple" else ("🔴 No cumple" if estado == "No cumple" else "⚪ No evaluable")
        comp = _celda_es_vs_pais(parametro, pais_fila, unidad_reg, unidad_es)
        # Límite normativo compacto (en la misma fila).
        if _sin_limite(inf) and _sin_limite(sup):
            limite_cell = "Sin límite numérico"
        else:
            limite_cell = f"{inf} / {sup}" + (f" {unidad_reg}" if unidad_reg else "")
        # Valor evaluado, con nota de conversión si la hubo.
        valor_eval = item.get("valor_evaluado", valor)
        valor_usr = item.get("valor_usuario", valor)
        unidad_usr = item.get("unidad_usuario", unidad or "")
        conv = item.get("conversion", "")
        if conv and str(valor_usr) != str(valor_eval):
            celda = f"{valor_eval} {unidad_reg} (de {valor_usr} {unidad_usr})"
            if conv not in conversiones and "Sin conversión" not in conv:
                conversiones.append(conv)
        else:
            celda = f"{valor_eval} {unidad_reg}".strip()
        lines.append(f"| {pais_fila} | {nombre} | {celda} | {limite_cell} | {condiciones} | {res} | {detalle} | {comp} |")
        evidencias.append(f"- **{pais_fila}** · {nombre}: {_cita_oficial(item)}")
        if item.get("discrepancia"):
            evidencias.append(f"  - ⚠ Discrepancia con el Excel (prevalece la fuente oficial): {item['discrepancia']}")
        if item.get("nota"):
            evidencias.append(f"  - 📝 Nota de la fuente: {item['nota']}")
        if estado == "Cumple":
            cumple_en.append(f"{pais_fila} ({nombre})")
    bloques = list(lines)
    if conversiones:
        bloques += ["", "**Conversión aplicada**", ""] + [f"- {c}" for c in conversiones]
    if evidencias:
        bloques += ["", "**Evidencias**", ""] + evidencias
    if todos:
        resumen = ", ".join(cumple_en) if cumple_en else "ninguno de los evaluados"
        bloques += ["", f"**Cumple en:** {resumen}."]
    return "\n".join(bloques)


def _comparar_normativa(parametro: str, paises: list) -> str:
    """Compara la NORMATIVA (límites/unidades) entre países, SIN valor del usuario.

    Muestra una fila por país enfrentando sus límites, con la columna de
    comparabilidad cruzada (España como referencia).
    """
    unidad_es = _unidad_de_pais(parametro, PAIS_BASE)
    if len(paises) <= 1:
        titulo = f"**Límites normativos de {parametro} — {paises[0] if paises else ''}**"
    else:
        titulo = f"**Comparación normativa de {parametro} — {' vs '.join(paises)}**"
    lines = [
        titulo,
        "",
        "| País | Parámetro | Límites | Unidad | Condiciones | Comparabilidad normativa |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    hubo = False
    estados: list = []
    normalizaciones: list = []  # PCS/Wobbe llevados a condiciones de España (Tabla A.1)
    evidencias: list = []
    for pais in paises:
        try:
            resp = _consultar_norma(parametro, pais)
        except Exception:
            continue
        for m in resp.get("matches", []):
            hubo = True
            nombre = str(m.get("parametro") or parametro).strip()
            inf = _txt(m.get("limite_inferior")) or "-"
            sup = _txt(m.get("limite_superior")) or "-"
            unidad_reg = _txt(m.get("unidad")).strip("()").replace("^3", "³").replace("^2", "²")
            cond = _normalize_condition_text(_txt(m.get("condiciones"))) or "—"
            if _sin_limite(inf) and _sin_limite(sup):
                limite = "Sin límite numérico"
            else:
                limite = f"{inf} / {sup}"
            comp = _celda_es_vs_pais(parametro, pais, unidad_reg, unidad_es)
            if _norm_pais(pais) != _norm_pais(PAIS_BASE):
                estados.append((pais, _estado_comparabilidad(parametro, unidad_es, unidad_reg)))
                nota = _limites_en_condiciones_espana(parametro, pais, inf, sup, unidad_reg, nombre)
                if nota:
                    normalizaciones.append(nota)
            lines.append(f"| {pais} | {nombre} | {limite} | {unidad_reg or '—'} | {cond} | {comp} |")
            evidencias.append(f"- **{pais}** · {nombre}: {_cita_oficial(m)}")
            if m.get("discrepancia"):
                evidencias.append(f"  - ⚠ Discrepancia con el Excel (prevalece la oficial): {m['discrepancia']}")
            if m.get("nota"):
                evidencias.append(f"  - 📝 Nota de la fuente: {m['nota']}")
    if not hubo:
        return f"No encontré datos normativos de '{parametro}' en {', '.join(paises)}."
    if normalizaciones:
        lines += [
            "",
            "**Normalización a condiciones de España** (combustión 0 °C; ISO 13443, Tabla A.1)",
            "",
        ] + normalizaciones
    if evidencias:
        lines += ["", "**Evidencias (fuente oficial)**", ""] + evidencias
    if estados:
        estados_u = list(dict.fromkeys(estados))  # dedup (un país puede tener varias filas)
        sint = "; ".join(f"España vs {p}: {e}" for p, e in estados_u)
        lines += ["", f"**Síntesis:** {sint}."]
    return "\n".join(lines)


def _num_limite(x: Any) -> Optional[float]:
    """Extrae el primer número de un límite ('48,17', 'no especificado'…)."""
    s = str(x).strip().replace(",", ".")
    m = re.search(r"[-+]?[0-9]*\.?[0-9]+", s)
    return float(m.group(0)) if m else None


def _limites_en_condiciones_espana(
    parametro: str, pais: str, inf: Any, sup: Any, unidad_reg: str, nombre: str
) -> Optional[str]:
    """Para PCS/Wobbe de un país no-España: devuelve sus límites llevados a las
    condiciones españolas (unidad kWh/m³ + combustión 0 °C, Tabla A.1). None si no aplica."""
    if _slug_param_comb(parametro) not in {"pcs", "wobbe"}:
        return None
    cond = _COND_PAIS.get(_norm_pais(pais))
    misma_unidad = _normalize_unit(unidad_reg or "kwh/m3") == _normalize_unit("kWh/m³")
    # Si el país ya está en base española (combustión 0 °C) y misma unidad, no hay nada que cambiar.
    if cond == (0, 0) and misma_unidad:
        return None
    vi, vs = _num_limite(inf), _num_limite(sup)
    if vi is None and vs is None:
        return None
    factor = 1.0
    partes = []
    for v in (vi, vs):
        if v is None:
            partes.append("—")
            continue
        # 1) Unidad -> kWh/m³ (si hace falta), de forma determinista.
        if unidad_reg and not misma_unidad:
            cu = convertir_unidades(v, unidad_reg, "kWh/m³", parametro)
            if "valor_convertido" in cu:
                v = cu["valor_convertido"]
        # 2) Condición de combustión del país -> España (Tabla A.1).
        cr = convertir_a_condiciones_espana(v, parametro, pais)
        factor = cr.get("factor", 1.0)
        partes.append(f"{cr['valor_convertido']:.2f}".replace(".", ","))
    return f"- **{pais}** · {nombre}: {partes[0]} / {partes[1]} kWh/m³ (de {unidad_reg or 'kWh/m³'}, combustión {cond[0] if cond else '?'} → 0 °C, factor {factor:g})"


def _coma(x: Any) -> str:
    try:
        return f"{round(float(x), 4):g}".replace(".", ",")
    except (TypeError, ValueError):
        return str(x)


def _coma2(x: Any) -> str:
    """Como _coma pero a 2 decimales. Se usa para los valores DERIVADOS (conversión de
    unidad y normalización ISO 13443 a condiciones de España): a 4 decimales dan una falsa
    precisión (p. ej. 13,6516) cuyo último dígito depende del orden del cálculo. Los valores
    NATIVOS de cada norma se muestran con _coma (su precisión regulatoria)."""
    try:
        return f"{round(float(x), 2):g}".replace(".", ",")
    except (TypeError, ValueError):
        return str(x)


def _es_consulta_condiciones(texto_norm: str) -> bool:
    """¿Pide llevar un valor a las condiciones de referencia (combustión) de España?"""
    claves = (
        "condiciones de españa", "condiciones de espana", "condiciones espanola",
        "condiciones española", "condiciones espanolas", "condiciones españolas",
        "condicion de combustion", "condiciones de combustion", "temperatura de combustion",
        "tabla a1", "tabla a.1", "iso 13443", "base espanola", "base española",
        "referencia espanola", "referencia española", "a condiciones espanola",
        "normalizar a espana", "normalizar a españa", "pasar a espana", "pasar a españa",
    )
    return any(k in texto_norm for k in claves)


def _responder_condiciones(parametro: str, valor, unidad, pais_origen: str) -> str:
    """Convierte un PCS/Wobbe de `pais_origen` a las condiciones de combustión de España."""
    display = DISPLAY_MAP.get(parametro, parametro)
    if _slug_param_comb(parametro) not in {"pcs", "wobbe"}:
        return (
            f"El parámetro **{display}** se mide a 0 °C y no depende de la temperatura de "
            f"combustión, así que es directamente comparable entre {pais_origen} y España "
            "sin ajustar las condiciones de referencia."
        )
    if valor is None:
        # Sin valor del usuario: muestra los límites del país ya normalizados a España.
        return _comparar_normativa(parametro, [PAIS_BASE, pais_origen])

    pasos = []
    v = float(valor)
    u = unidad or ""
    if u and _normalize_unit(u) != _normalize_unit("kWh/m³"):
        cu = convertir_unidades(v, u, "kWh/m³", parametro)
        if "valor_convertido" in cu:
            pasos.append(f"- Unidad: {_coma(v)} {u} → {_coma(cu['valor_convertido'])} kWh/m³  ({cu.get('formula','')})")
            v = cu["valor_convertido"]
    cr = convertir_a_condiciones_espana(v, parametro, pais_origen)
    pasos.append(
        f"- Combustión: {cr['condiciones_origen']} → {cr['condiciones_destino']}  "
        f"(Tabla A.1, factor {cr['factor']:g})"
    )
    return "\n".join([
        f"**{display} de {pais_origen} en condiciones de España**",
        "",
        f"**{_coma(valor)} {unidad or ''}** ({pais_origen}) → **{_coma(cr['valor_convertido'])} kWh/m³** (España)",
        "",
        *pasos,
        "",
        f"_Fuente del factor: {cr['fuente']}._",
    ])


# --- Comparativa estructurada (sección de la web con desplegables) ----------
# Unidades ofrecidas por parámetro (la 1.ª es la base normativa española).
PARAMETROS_UI = [
    {"slug": "wobbe", "label": "Índice de Wobbe", "unidades": ["kWh/m³", "MJ/m³"]},
    {"slug": "pcs", "label": "PCS (Poder Calorífico Superior)", "unidades": ["kWh/m³", "MJ/m³"]},
    {"slug": "densidad relativa", "label": "Densidad relativa", "unidades": ["adimensional"]},
    {"slug": "s total", "label": "Azufre total (S)", "unidades": ["mg/m³", "mg/Nm³", "ppm"]},
    {"slug": "h2s+cos", "label": "H₂S + COS", "unidades": ["mg/m³", "mg/Nm³", "ppm"]},
    {"slug": "rsh", "label": "Mercaptanos (RSH)", "unidades": ["mg/m³", "mg/Nm³", "ppm"]},
    {"slug": "o2", "label": "O₂ (oxígeno)", "unidades": ["% molar", "ppm"]},
    {"slug": "co2", "label": "CO₂", "unidades": ["% molar", "ppm"]},
    {"slug": "h2o(rocío)", "label": "Punto de rocío del agua (H₂O)", "unidades": ["°C", "K", "°F"]},
    {"slug": "hc(rocío)", "label": "Punto de rocío de HC", "unidades": ["°C", "K", "°F"]},
]
# Catálogo de parámetros del BIOMETANO (dominio nuevo; NO altera el de gas natural).
# Cada entrada lleva su clave en la ontología (`onto`) para no tocar los mapas
# _PARAM_A_ONTO del gas natural (riesgo R2). Jurisdicción única: EN 16723-1.
PARAMETROS_UI_BIOMETANO = [
    {"slug": "ch4_min", "label": "CH₄ (metano) mínimo", "unidades": ["% molar"], "onto": "CH4_MIN"},
    {"slug": "o2", "label": "O₂ (oxígeno)", "unidades": ["% molar", "ppm"], "onto": "O2"},
    {"slug": "co2", "label": "CO₂", "unidades": ["% molar", "ppm"], "onto": "CO2"},
    {"slug": "co", "label": "CO (monóxido de carbono)", "unidades": ["% molar", "ppm"], "onto": "CO"},
    {"slug": "s total", "label": "Azufre total (S)", "unidades": ["mg/m³", "mg/Nm³", "ppm"], "onto": "S_TOTAL"},
    {"slug": "h2s+cos", "label": "H₂S + COS", "unidades": ["mg/m³", "mg/Nm³", "ppm"], "onto": "H2S_COS"},
    {"slug": "h2o(rocío)", "label": "Punto de rocío del agua (H₂O)", "unidades": ["°C", "K", "°F"], "onto": "PR_H2O"},
    {"slug": "siloxanos", "label": "Siloxanos (Si total)", "unidades": ["mg/m³"], "onto": "SILOXANOS"},
    {"slug": "comp_oil", "label": "Aceite de compresor", "unidades": ["mg/m³"], "onto": "COMP_OIL"},
    {"slug": "aminas", "label": "Aminas", "unidades": ["mg/m³", "ppm"], "onto": "AMINAS"},
    {"slug": "nh3", "label": "Amoníaco (NH₃)", "unidades": ["mg/m³", "ppm"], "onto": "NH3"},
    {"slug": "halogenados", "label": "Compuestos halogenados (Cl+F)", "unidades": ["mg/m³", "ppm"], "onto": "HALOGENADOS"},
]
# Catálogo del HIDRÓGENO (ISO 14687:2019 Grade D / EN 17124). Trazas en ppm (≡ μmol/mol);
# pureza en % molar; partículas en mg/kg. Jurisdicción única: ISO 14687 Grade D.
PARAMETROS_UI_HIDROGENO = [
    {"slug": "h2_pureza", "label": "Pureza H₂ (mínimo)", "unidades": ["% molar"], "onto": "H2_PUREZA"},
    {"slug": "no_h2", "label": "Total gases no-H₂", "unidades": ["ppm"], "onto": "NO_H2"},
    {"slug": "h2o", "label": "Agua (H₂O)", "unidades": ["ppm"], "onto": "H2O"},
    {"slug": "thc", "label": "Hidrocarburos (excepto CH₄)", "unidades": ["ppm"], "onto": "THC"},
    {"slug": "ch4", "label": "Metano (CH₄)", "unidades": ["ppm"], "onto": "CH4"},
    {"slug": "o2", "label": "O₂ (oxígeno)", "unidades": ["ppm"], "onto": "O2"},
    {"slug": "he", "label": "Helio (He)", "unidades": ["ppm"], "onto": "HE"},
    {"slug": "n2", "label": "Nitrógeno (N₂)", "unidades": ["ppm"], "onto": "N2"},
    {"slug": "ar", "label": "Argón (Ar)", "unidades": ["ppm"], "onto": "AR"},
    {"slug": "co2", "label": "CO₂", "unidades": ["ppm"], "onto": "CO2"},
    {"slug": "co", "label": "CO (monóxido de carbono)", "unidades": ["ppm"], "onto": "CO"},
    {"slug": "s total", "label": "Azufre total (S)", "unidades": ["ppm"], "onto": "S_TOTAL"},
    {"slug": "hcho", "label": "Formaldehído (HCHO)", "unidades": ["ppm"], "onto": "HCHO"},
    {"slug": "hcooh", "label": "Ácido fórmico (HCOOH)", "unidades": ["ppm"], "onto": "HCOOH"},
    {"slug": "nh3", "label": "Amoníaco (NH₃)", "unidades": ["ppm"], "onto": "NH3"},
    {"slug": "halogenados", "label": "Compuestos halogenados", "unidades": ["ppm"], "onto": "HALOGENADOS"},
    {"slug": "particulas", "label": "Partículas", "unidades": ["mg/kg"], "onto": "PARTICULAS"},
]
PAISES_UI = ["Portugal", "Francia", "Italia", "Alemania", "Países Bajos", "Bélgica", "Noruega", "Polonia", "Dinamarca", "Hungría", "Austria", "Suiza", "Chequia", "Grecia", "Irlanda", "Rumanía", "Eslovaquia", "Turquía", "Reino Unido", "UE"]  # España es siempre la base de referencia


def _es_mg_por_volumen(unidad: str) -> bool:
    """True si la unidad es una concentración másica por volumen (mg/m³, mg/Nm³, mg/Sm³).

    Estas dependen de la temperatura del volumen de referencia (gas ideal); el % mol,
    el ppm (molar) y las adimensionales NO.
    """
    s = (unidad or "").lower().replace(" ", "")
    return "mg" in s and ("m³" in s or "m3" in s or "nm" in s or "sm" in s)


def comparar_estructurado(parametro_slug: str, paises: list, unidad_destino: str = "") -> Dict[str, Any]:
    """Comparativa para la sección con desplegables. Devuelve filas estructuradas y,
    para PCS/Wobbe de países con distinta temperatura de combustión, el equivalente
    en condiciones de España (ISO 13443, Tabla A.1)."""
    orden = [PAIS_BASE] + [p for p in (paises or []) if _norm_pais(p) != _norm_pais(PAIS_BASE)]
    display = DISPLAY_MAP.get(parametro_slug, parametro_slug)
    es_combustion = _slug_param_comb(parametro_slug) in {"pcs", "wobbe"}
    filas: list = []
    notas: list = []
    for pais in orden:
        try:
            matches = _consultar_norma(parametro_slug, pais).get("matches", [])
        except Exception:
            matches = []
        for m in matches:
            nombre = str(m.get("parametro") or parametro_slug).strip()
            inf_raw = _txt(m.get("limite_inferior")) or "-"
            sup_raw = _txt(m.get("limite_superior")) or "-"
            unidad_reg = _txt(m.get("unidad")).strip("()").replace("^3", "³").replace("^2", "²")
            # Condiciones POR PARÁMETRO (de la notación del registro), no del país:
            # en la UE el Wobbe es 15/15 pero las concentraciones se expresan a 0/0.
            ccomb, cmed = _parse_notac(m.get("notacion") or "(0/0)")
            cond_txt = f"comb. {_coma(ccomb)} °C · med. {_coma(cmed)} °C"
            if "m3" in _normalize_unit(unidad_reg):  # unidad por volumen de gas → aclarar Nm³/Sm³
                cond_txt += " (Nm³)" if cmed == 0 else (" (Sm³)" if cmed == 15 else "")
            sin_lim = _sin_limite(inf_raw) and _sin_limite(sup_raw)
            es_base = _norm_pais(pais) == _norm_pais(PAIS_BASE)

            def a_unidad(raw):
                v = _num_limite(raw)
                if v is None:
                    return None
                if unidad_destino and unidad_destino != "adimensional" and unidad_reg \
                        and _normalize_unit(unidad_destino) != _normalize_unit(unidad_reg):
                    cu = convertir_unidades(v, unidad_reg, unidad_destino, parametro_slug)
                    return cu.get("valor_convertido") if "valor_convertido" in cu else None
                return v

            vi, vs = a_unidad(inf_raw), a_unidad(sup_raw)
            unidad_out = (unidad_destino or unidad_reg or "—")

            vi_es = vs_es = None
            factor = 1.0
            if not es_base and not sin_lim:
                if es_combustion:
                    if vi is not None:
                        vi_es, factor = _a_espana_por_registro(vi, parametro_slug, pais, m)
                    if vs is not None:
                        vs_es, factor = _a_espana_por_registro(vs, parametro_slug, pais, m)
                    if factor and factor != 1.0:
                        notas.append(f"{pais}: combustión {_coma(ccomb)} °C → 0 °C (× {factor:g}, ISO 13443 Tabla A.1)")
                elif cmed and cmed != 0 and _es_mg_por_volumen(unidad_out):
                    # Concentración MÁSICA (mg/m³) referida a un volumen a T ≠ 0 °C
                    # (p. ej. Italia: mg/Sm³ a 15 °C). Se normaliza el volumen a 0 °C
                    # con factor (273,15 + T_vol)/273,15 (base de gas ideal). El % mol,
                    # ppm (molar) y adimensionales NO dependen de la temperatura.
                    fvol = (273.15 + cmed) / 273.15
                    vi_es = round(vi * fvol, 6) if vi is not None else None
                    vs_es = round(vs * fvol, 6) if vs is not None else None
                    factor = fvol
                    notas.append(f"{pais}: volumen {_coma(cmed)} °C → 0 °C (× {fvol:.4f}, base de gas ideal)")
                else:
                    # No depende de la temperatura de combustión ni del volumen: mismo valor, referido a 0/0.
                    vi_es, vs_es = vi, vs

            # ¿El límite mostrado es un valor DERIVADO (hubo conversión de unidad)? Entonces
            # se muestra a 2 decimales (evita la falsa precisión); si es el valor nativo de la
            # norma, se muestra con su precisión (_coma). La normalización a España es SIEMPRE
            # derivada → 2 decimales.
            hubo_conversion = bool(
                unidad_destino and unidad_destino != "adimensional" and unidad_reg
                and _normalize_unit(unidad_destino) != _normalize_unit(unidad_reg)
            )

            def rng(a, b, dec=4):
                if sin_lim:
                    return "Sin límite numérico"
                fmt = _coma2 if dec == 2 else _coma
                return f"{fmt(a) if a is not None else '—'} / {fmt(b) if b is not None else '—'}"

            filas.append({
                "pais": pais,
                "parametro": nombre,
                "limite": rng(vi, vs, 2 if hubo_conversion else 4),
                "unidad": "—" if sin_lim else unidad_out,
                "condiciones": cond_txt,
                "es_base": es_base,
                "limite_espana": (
                    None if es_base
                    else ("Sin límite numérico" if sin_lim
                          else (rng(vi_es, vs_es, 2) if (vi_es is not None or vs_es is not None) else None))
                ),
                "factor_iso": (factor if (es_combustion and not es_base and factor != 1.0) else None),
                # Cita de la fuente oficial (para mostrar en el frontend).
                "fuente": m.get("documento") or "",
                "organismo": m.get("organismo") or "",
                "fecha": m.get("fecha") or "",
                "articulo": m.get("articulo") or "",
                "url": m.get("url") or m.get("pdf") or "",
                "estado": m.get("estado") or "",
                "discrepancia": m.get("discrepancia") or "",
                # Condiciones/matices que el reglamento adjunta al límite (p. ej. el O₂
                # de la EN 16726: ≤1 % general, ≤0,01 %/0,001 % por evaluación).
                "nota": m.get("nota") or "",
            })
    return {
        "parametro": display,
        "unidad": unidad_destino,
        "es_combustion": es_combustion,
        "filas": filas,
        "notas_iso": list(dict.fromkeys(notas)),
    }


# --- Matriz comparativa avanzada (heatmap) ----------------------------------
# Filas = países, columnas = parámetros. Cada celda: rango normalizado a condiciones
# de España, un NIVEL frente a España (para el color) y un FLAG si hubo diferencia
# metodológica (unidad o condiciones distintas → se aplicó conversión).
PAISES_MATRIZ = ["España", "Portugal", "Francia", "Italia", "Alemania", "Países Bajos", "Bélgica", "Noruega", "Polonia", "Dinamarca", "Hungría", "Austria", "Suiza", "Chequia", "Grecia", "Irlanda", "Rumanía", "Eslovaquia", "Turquía", "Reino Unido", "UE"]

# Enrutado por tipo de gas (para "Analizar gas"). El gas natural es el actual.
# El biometano usa su propio catálogo y una única jurisdicción (EN 16723-1); NO se
# mezcla con la matriz de 21 países del gas natural (riesgos R4/R5).
CATALOGO_POR_GAS = {"gas_natural": PARAMETROS_UI, "biometano": PARAMETROS_UI_BIOMETANO,
                    "hidrogeno": PARAMETROS_UI_HIDROGENO}
# Las tres comparativas usan las MISMAS 4 jurisdicciones base (como empezó el gas natural).
# Biometano: compartidos = spec nacional de gas; específicos = EN 16723-1 (UE) / GRTgaz (FR).
# Hidrógeno: ISO 14687 (producto) aplicable en las 4 (aún no hay spec nacional/UE de red).
JURISDICCIONES_POR_GAS = {"gas_natural": PAISES_MATRIZ,
                          "biometano": ["España", "Portugal", "Francia", "UE"],
                          "hidrogeno": ["España", "Portugal", "Francia", "UE"]}
# Nombre legible de cada código de jurisdicción para la cabecera de la tabla.
JURISDICCION_DISPLAY = {"España": "España", "Portugal": "Portugal", "Francia": "Francia", "UE": "UE"}


def _celda_heatmap(slug, pais, unidad_es, notac_es, es_rng, es_maximo, ancho_es):
    es_base = _norm_pais(pais) == _norm_pais(PAIS_BASE)
    resp = fuente_oficial.consultar(slug, pais)
    if not resp.get("matches"):
        return {"valor": "—", "nivel": "sin_dato", "flag": False, "flag_desc": ""}
    m = resp["matches"][0]
    r = _rango_de_match(m, slug, pais, unidad_es)
    est = r.get("estado")

    def fr(lo, hi):
        return f"{_coma(lo) if lo is not None else '—'} / {_coma(hi) if hi is not None else '—'}"

    if est in ("sin_datos", "sin_dato"):
        return {"valor": "—", "nivel": "sin_dato", "flag": False, "flag_desc": ""}
    if est == "incomparable":
        return {"valor": "≠ unidades", "nivel": "incomparable", "flag": True, "flag_desc": "No comparable de forma determinista"}
    if est == "sin_limite":
        return {"valor": "Sin límite", "nivel": "sin_limite", "flag": False, "flag_desc": ""}

    # Flag metodológico: solo si hay límite real con unidad o condiciones RELEVANTES
    # distintas a España (la combustión solo cuenta para PCS/Wobbe).
    difs = []
    if not es_base:
        u = _txt(m.get("unidad"))
        if u and unidad_es and _normalize_unit(u) != _normalize_unit(unidad_es):
            difs.append(f"unidad {u}")
        if m.get("notacion") and notac_es and not _condiciones_iguales(slug, m.get("notacion"), notac_es):
            difs.append(f"condiciones {m.get('notacion')}")
    flag = bool(difs)
    flag_desc = (" · ".join(difs) + " → convertido a condiciones de España") if flag else ""

    if es_base:
        nivel = "base"
    elif es_rng.get("estado") != "ok":
        nivel = "sin_ref"
    else:
        lo = r["lo"] if r["lo"] is not None else float("-inf")
        hi = r["hi"] if r["hi"] is not None else float("inf")
        if es_maximo:
            es_hi = es_rng["hi"] if es_rng["hi"] is not None else float("inf")
            nivel = "restrictivo" if hi < es_hi - 1e-9 else ("amplio" if hi > es_hi + 1e-9 else "igual")
        else:
            ancho = hi - lo
            if ancho_es is None:
                nivel = "igual"
            else:
                nivel = "restrictivo" if ancho < ancho_es - 1e-9 else ("amplio" if ancho > ancho_es + 1e-9 else "igual")
    return {"valor": fr(r["lo"], r["hi"]), "nivel": nivel, "flag": flag, "flag_desc": flag_desc}


def matriz_comparativa() -> Dict[str, Any]:
    """Heatmap: filas=países, columnas=parámetros, celdas con nivel (color) y flag."""
    cols = []
    filas = {p: {} for p in PAISES_MATRIZ}
    for pinfo in PARAMETROS_UI:
        slug = pinfo["slug"]
        es = fuente_oficial.consultar(slug, PAIS_BASE)
        esm = es["matches"][0] if es.get("matches") else None
        unidad_es = (esm.get("unidad") if esm else "") or ""
        # La unidad (la de España, a la que se normaliza todo) se muestra en la cabecera.
        cols.append({"slug": slug, "label": pinfo["label"], "unidad": unidad_es})
        notac_es = esm.get("notacion") if esm else None
        es_rng = _rango_de_match(esm, slug, PAIS_BASE, unidad_es) if esm else {"estado": "sin_datos"}
        es_maximo = es_rng.get("estado") == "ok" and es_rng.get("lo") is None
        ancho_es = None
        if es_rng.get("estado") == "ok":
            lo = es_rng["lo"] if es_rng["lo"] is not None else float("-inf")
            hi = es_rng["hi"] if es_rng["hi"] is not None else float("inf")
            ancho_es = hi - lo
        for pais in PAISES_MATRIZ:
            filas[pais][slug] = _celda_heatmap(slug, pais, unidad_es, notac_es, es_rng, es_maximo, ancho_es)
    return {
        "paises": PAISES_MATRIZ,
        "parametros": cols,
        "filas": [{"pais": p, "celdas": filas[p]} for p in PAISES_MATRIZ],
        "unidad_nota": "Valores normalizados a unidad y condiciones de España (ISO 13443 para PCS/Wobbe).",
    }


# --- Fuente normativa: ¿de qué reglamento procede cada dato? ----------------
# Se lee de la ontología validada (data/ontologia/ontologia_enagas.yaml), que
# guarda la fuente, el artículo y la página verificados por parámetro y país.
_ONTOLOGIA_CACHE: Dict[str, Any] = {}
_PARAM_A_ONTO = {
    "wobbe": "WOBBE", "pcs": "PCS", "densidad relativa": "DENS_REL",
    "s total": "S_TOTAL", "h2s+cos": "H2S_COS", "rsh": "RSH",
    "o2": "O2", "co2": "CO2", "h2o(rocío)": "PR_H2O", "hc(rocío)": "PR_HC",
}
_PAIS_A_CODIGO = {"espana": "ES", "portugal": "PT", "francia": "FR", "ue": "UE", "europa": "UE",
                  "italia": "IT", "italy": "IT", "alemania": "DE", "germany": "DE",
                  "deutschland": "DE", "germania": "DE",
                  "paises bajos": "NL", "holanda": "NL", "netherlands": "NL", "nederland": "NL",
                  "belgica": "BE", "belgium": "BE", "belgique": "BE", "belgie": "BE",
                  "noruega": "NOR", "norway": "NOR", "norge": "NOR",
                  "polonia": "PL", "poland": "PL", "polska": "PL",
                  "dinamarca": "DK", "denmark": "DK", "danmark": "DK",
                  "hungria": "HU", "hungary": "HU", "magyarorszag": "HU",
                  "austria": "AT", "osterreich": "AT",
                  "suiza": "CH", "switzerland": "CH", "schweiz": "CH", "suisse": "CH",
                  "chequia": "CZ", "republica checa": "CZ", "czech": "CZ", "czechia": "CZ", "cesko": "CZ",
                  "grecia": "GR", "greece": "GR", "hellas": "GR",
                  "irlanda": "IE", "ireland": "IE", "eire": "IE",
                  "rumania": "RO", "romania": "RO",
                  "eslovaquia": "SK", "slovakia": "SK", "slovensko": "SK",
                  "turquia": "TR", "turkey": "TR", "turkiye": "TR",
                  "reino unido": "GB", "united kingdom": "GB", "gran bretana": "GB", "britain": "GB"}
_CODIGO_A_PAIS = {"ES": "España", "PT": "Portugal", "FR": "Francia", "UE": "UE",
                  "IT": "Italia", "DE": "Alemania", "NL": "Países Bajos", "BE": "Bélgica",
                  "NOR": "Noruega", "PL": "Polonia", "DK": "Dinamarca", "HU": "Hungría", "AT": "Austria", "CH": "Suiza", "CZ": "Chequia", "GR": "Grecia",
                  "IE": "Irlanda", "RO": "Rumanía", "SK": "Eslovaquia", "TR": "Turquía", "GB": "Reino Unido"}
_PAISES_FUENTE = ["España", "Portugal", "Francia", "Italia", "Alemania", "Países Bajos", "Bélgica", "Noruega", "Polonia", "Dinamarca", "Hungría", "Austria", "Suiza", "Chequia", "Grecia", "Irlanda", "Rumanía", "Eslovaquia", "Turquía", "Reino Unido", "UE"]


def _cargar_ontologia() -> Dict[str, Any]:
    """Carga (una sola vez) la ontología validada. Devuelve {} si no es posible."""
    if "data" in _ONTOLOGIA_CACHE:
        return _ONTOLOGIA_CACHE["data"]
    data: Dict[str, Any] = {}
    try:
        import yaml  # en requirements.txt
        ruta = os.path.join(os.path.dirname(__file__), "data", "ontologia", "ontologia_enagas.yaml")
        if not os.path.exists(ruta):
            import glob
            cand = glob.glob(os.path.join(os.path.dirname(__file__), "data", "**", "ontologia_enagas.yaml"), recursive=True)
            ruta = cand[0] if cand else ruta
        with open(ruta, encoding="utf-8") as fh:
            data = yaml.safe_load(fh) or {}
    except Exception as exc:  # noqa: BLE001
        print(f"[fuente] No se pudo cargar la ontología ({exc}).")
    _ONTOLOGIA_CACHE["data"] = data
    return data


def _glosario_fuentes() -> Dict[str, Dict[str, Any]]:
    onto = _cargar_ontologia()
    fuentes = (onto.get("ontologia") or {}).get("fuentes_normativas") or []
    return {f.get("id"): f for f in fuentes if isinstance(f, dict) and f.get("id")}


def _fuentes_de(parametro_slug: str, paises: list) -> list:
    """Lista de fuentes [{pais, reglamento, ubicacion, estado, publicacion, nota}]."""
    onto = _cargar_ontologia()
    params = onto.get("parametros") or {}
    clave = _PARAM_A_ONTO.get(parametro_slug)
    if not clave or clave not in params:
        return []
    limites = params[clave].get("limites") or {}
    glos = _glosario_fuentes()
    salida = []
    for pais in paises:
        cod = _PAIS_A_CODIGO.get(_norm_pais(pais))
        if not cod or cod not in limites:
            continue
        lim = limites[cod] or {}
        info = glos.get(lim.get("fuente"), {})
        salida.append({
            "pais": _CODIGO_A_PAIS.get(cod, pais),
            "reglamento": info.get("nombre") or lim.get("fuente") or "No consta en la fuente",
            "ubicacion": lim.get("articulo") or info.get("tabla_calidad") or "—",
            "estado": lim.get("estado_verificacion") or "—",
            "publicacion": info.get("publicacion") or info.get("referencia_legal") or "",
            "nota": lim.get("nota") or "",
        })
    return salida


def _responder_fuente(parametro_slug: str, paises: list) -> str:
    datos = _fuentes_de(parametro_slug, paises)
    display = DISPLAY_MAP.get(parametro_slug, parametro_slug)
    if not datos:
        return (
            f"No tengo registrada la fuente normativa de '{display}' en la ontología validada "
            f"para {', '.join(paises)}."
        )
    lines = [
        f"**Fuente normativa de {display}**",
        "",
        "| País | Reglamento / Norma | Ubicación en el documento | Estado |",
        "| --- | --- | --- | --- |",
    ]
    extra = []
    for d in datos:
        lines.append(f"| {d['pais']} | {d['reglamento']} | {d['ubicacion']} | {d['estado']} |")
        if d["publicacion"]:
            extra.append(f"- **{d['pais']}**: {d['publicacion']}")
        if d["nota"]:
            extra.append(f"- **{d['pais']}** (nota): {d['nota']}")
    if extra:
        lines += ["", "**Publicación / referencia**", ""] + extra
    return "\n".join(lines)


def _listar_fuentes() -> str:
    """Lista los reglamentos primarios realmente usados por los parámetros."""
    onto = _cargar_ontologia()
    params = onto.get("parametros") or {}
    glos = _glosario_fuentes()
    usados: list = []
    for p in params.values():
        for lim in (p.get("limites") or {}).values():
            fc = (lim or {}).get("fuente")
            if fc and fc not in usados:
                usados.append(fc)
    if not usados:
        return "No pude cargar la lista de reglamentos de la ontología validada."
    lines = [
        "**Reglamentos y normas que utiliza el comparador**",
        "",
        "| País | Reglamento / Norma | Publicación |",
        "| --- | --- | --- |",
    ]
    for fc in usados:
        info = glos.get(fc, {})
        pais = _CODIGO_A_PAIS.get(info.get("pais"), info.get("pais") or "—")
        nombre = info.get("nombre") or fc
        pub = info.get("publicacion") or info.get("referencia_legal") or "—"
        lines.append(f"| {pais} | {nombre} | {pub} |")
    return "\n".join(lines)


def _normativa_de_pais(paises: list) -> str:
    """#1 — «¿Qué normativa aplican en [país]?» Lista las normas oficiales que ese país
    aplica a la calidad del gas (fuentes distintas usadas en sus parámetros)."""
    onto = _cargar_ontologia()
    params = onto.get("parametros") or {}
    glos = _glosario_fuentes()
    bloques: list = []
    for pais in paises:
        cod = _PAIS_A_CODIGO.get(_norm_pais(pais))
        if not cod:
            continue
        usados: list = []
        for p in params.values():
            fc = ((p.get("limites") or {}).get(cod) or {}).get("fuente")
            if fc and fc not in usados:
                usados.append(fc)
        nombre_pais = _CODIGO_A_PAIS.get(cod, pais)
        if not usados:
            bloques.append(f"No tengo registrada la normativa de calidad del gas de {nombre_pais}.")
            continue
        lineas = [f"**Normativa de calidad del gas aplicable en {nombre_pais}**", ""]
        for fc in usados:
            info = glos.get(fc, {})
            partes = [f"- **{info.get('nombre') or fc}**"]
            if info.get("organismo"):
                partes.append(info["organismo"])
            pub = info.get("publicacion") or info.get("referencia_legal") or ""
            if pub:
                partes.append(pub)
            linea = " · ".join(partes)
            url = fuente_oficial.url_de(info)
            if url:
                linea += f" — {url}"
            lineas.append(linea)
            ambito = info.get("tabla_calidad") or info.get("ambito") or ""
            if ambito:
                lineas.append(f"  - {ambito}")
        bloques.append("\n".join(lineas))
    return "\n\n".join(bloques) if bloques else "No encontré normativa para el país indicado."


def _a_espana_por_registro(v: float, parametro: str, pais: str, m: Optional[Dict[str, Any]]) -> tuple:
    """Convierte `v` (PCS/Wobbe) a las condiciones de España (0/0) usando las condiciones
    del REGISTRO (su `notacion`) si las tiene; si no, las del país. Necesario cuando un país
    tiene parámetros en condiciones distintas (p. ej. Alemania: Wobbe/CO₂ @25/0 pero PCS
    @15/15, de la lista UE 2021/C 78/05). Devuelve (valor_convertido, factor)."""
    notac = (m or {}).get("notacion")
    if notac:
        ncomb, nmed = _parse_notac(notac)
        cr = convertir_condiciones_referencia(v, parametro, ncomb, nmed, 0.0, 0.0)
    else:
        cr = convertir_a_condiciones_espana(v, parametro, pais)
    return cr.get("valor_convertido", v), cr.get("factor", 1.0)


def _rango_de_match(m: Dict[str, Any], parametro: str, pais: str, unidad_es: Optional[str]) -> Dict[str, Any]:
    """Lleva el rango de un registro `m` a la UNIDAD y CONDICIONES de España (ISO 13443)."""
    inf = _num_limite(_txt(m.get("limite_inferior")))
    sup = _num_limite(_txt(m.get("limite_superior")))
    if inf is None and sup is None:
        return {"estado": "sin_limite"}
    ureg = _txt(m.get("unidad")).strip("()").replace("^3", "³").replace("^2", "²")
    es_combustion = _slug_param_comb(parametro) in {"pcs", "wobbe"}
    base_es = _norm_pais(pais) == _norm_pais(PAIS_BASE)

    def conv(v):
        if v is None:
            return None
        if unidad_es and ureg and _normalize_unit(unidad_es) != _normalize_unit(ureg):
            c = convertir_unidades(v, ureg, unidad_es, parametro)
            if "valor_convertido" not in c:
                return "incomparable"
            v = c["valor_convertido"]
        if es_combustion and not base_es:
            v, _ = _a_espana_por_registro(v, parametro, pais, m)
        elif not base_es:
            # Concentración másica (mg/m³) a volumen ≠ 0 °C (p. ej. Italia, Sm³ a 15 °C):
            # normaliza el volumen a 0 °C (gas ideal). El % mol/ppm/adimensional no cambia.
            _cc, cmed = _parse_notac(m.get("notacion") or "(0/0)")
            if cmed and cmed != 0 and _es_mg_por_volumen(unidad_es or ureg):
                v = round(v * (273.15 + cmed) / 273.15, 6)
        return v

    lo, hi = conv(inf), conv(sup)
    if lo == "incomparable" or hi == "incomparable":
        return {"estado": "incomparable"}
    return {"estado": "ok", "lo": lo, "hi": hi}


def _rango_en_condiciones_es(parametro: str, pais: str, unidad_es: Optional[str]) -> Dict[str, Any]:
    """Rango de límites de `pais` para `parametro`, llevado a la UNIDAD y CONDICIONES de
    España. Devuelve {estado, lo, hi} con estado: 'ok' | 'sin_limite' | 'sin_datos' | 'incomparable'."""
    resp = _consultar_norma(parametro, pais)
    if not resp.get("matches"):
        return {"estado": "sin_datos"}
    return _rango_de_match(resp["matches"][0], parametro, pais, unidad_es)


def _intercambiabilidad(parametro: str, paises: Optional[list] = None) -> str:
    """#2 — «Teniendo en cuenta los límites de España de [parámetro], ¿con qué país
    podría intercambiar gas?» Compara el rango español con el de cada país (en
    condiciones de España) y decide si son intercambiables (solape de rangos)."""
    paises = paises or [p for p in ALL_COUNTRIES if _norm_pais(p) != _norm_pais(PAIS_BASE)]
    display = DISPLAY_MAP.get(parametro, parametro)
    unidad_es = _unidad_de_pais(parametro, PAIS_BASE)
    es = _rango_en_condiciones_es(parametro, PAIS_BASE, unidad_es)
    if es.get("estado") != "ok":
        return f"No tengo un límite numérico de España para {display}, así que no puedo evaluar la intercambiabilidad."
    es_lo = es["lo"] if es["lo"] is not None else float("-inf")
    es_hi = es["hi"] if es["hi"] is not None else float("inf")

    def fr(lo, hi):
        a = _coma(lo) if lo is not None else "—"
        b = _coma(hi) if hi is not None else "—"
        return f"{a} / {b}"

    lines = [
        f"**Intercambiabilidad de gas según {display}** (base España: {fr(es['lo'], es['hi'])} {unidad_es or ''})",
        "",
        "| País | Límite (en condiciones de España) | ¿Intercambiable? | Detalle |",
        "| --- | --- | --- | --- |",
    ]
    compatibles = []
    for pais in paises:
        r = _rango_en_condiciones_es(parametro, pais, unidad_es)
        est = r.get("estado")
        if est == "sin_datos":
            lines.append(f"| {pais} | — | ⚪ Sin datos | No consta el parámetro en su normativa |")
            continue
        if est == "incomparable":
            lines.append(f"| {pais} | (no convertible) | 🔴 No | Unidades/magnitud no comparables con España |")
            continue
        if est == "sin_limite":
            lines.append(f"| {pais} | Sin límite | 🟢 Sí | {pais} no fija este parámetro: admite el gas español |")
            compatibles.append(pais)
            continue
        lo = r["lo"] if r["lo"] is not None else float("-inf")
        hi = r["hi"] if r["hi"] is not None else float("inf")
        solapan = max(es_lo, lo) <= min(es_hi, hi) + 1e-9
        es_subset = (lo - 1e-9) <= es_lo and es_hi <= (hi + 1e-9)
        if not solapan:
            detalle = "Rangos sin solape: el gas español no cumpliría su norma"
            verdict = "🔴 No"
        elif es_subset:
            detalle = "Todo gas que cumple en España cumple también aquí"
            verdict = "🟢 Sí"
            compatibles.append(pais)
        else:
            detalle = "Solape parcial: parte del gas español cumple (revisar valor concreto)"
            verdict = "🟡 Parcial"
            compatibles.append(pais + " (parcial)")
        lines.append(f"| {pais} | {fr(r['lo'], r['hi'])} {unidad_es or ''} | {verdict} | {detalle} |")

    resumen = ", ".join(compatibles) if compatibles else "ninguno de los evaluados"
    lines += ["", f"**Intercambio posible con:** {resumen}.",
              "", "_Nota: evaluación por solape de límites del parámetro indicado; el intercambio real exige cumplir TODOS los parámetros a la vez._"]
    return "\n".join(lines)


# ============================================================================
# ALERTAS DE INCOMPATIBILIDAD EN CADENA (interconexión)
# ---------------------------------------------------------------------------
# Cuando el usuario pregunta por una interconexión (p. ej. "España-Francia-
# Alemania"), un gas debe cumplir el límite de TODOS los países a la vez. Por
# cada parámetro se calcula la INTERSECCIÓN de rangos (normalizados a España) y
# qué país impone la restricción más dura (cuello de botella). Si para algún
# parámetro no queda rango común → incompatibilidad. Reutiliza el motor
# comparativo (`_rango_en_condiciones_es`).
# ============================================================================

_KEYWORDS_INTERCONEXION = (
    "interconex", "cadena", "corredor", "transit", "pasa por", "pasando por",
    "circular por", "cuello de botella", "gasoducto", "atravesar", "atraviesa", "encadenad",
)


def _extraer_cadena_paises(mensaje: str) -> List[str]:
    """Países mencionados en el mensaje, EN ORDEN de aparición y sin repetir."""
    txt = _norm_pais(mensaje)
    hallados: List[tuple] = []
    for alias, cod in _PAIS_A_CODIGO.items():
        desde = 0
        while True:
            i = txt.find(alias, desde)
            if i < 0:
                break
            antes = txt[i - 1] if i > 0 else " "
            despues = txt[i + len(alias)] if i + len(alias) < len(txt) else " "
            if not antes.isalpha() and not despues.isalpha():  # palabra completa
                hallados.append((i, cod))
            desde = i + 1
    hallados.sort(key=lambda x: x[0])
    orden: List[str] = []
    for _, cod in hallados:
        pais = "España" if cod == "ES" else _CODIGO_A_PAIS.get(cod, cod)
        if pais not in orden:
            orden.append(pais)
    return orden


def _es_consulta_interconexion(mensaje: str, paises: List[str]) -> bool:
    """¿El mensaje pide un análisis de cadena/interconexión? (2+ países + señal explícita)."""
    if len(paises) < 2:
        return False
    t = _norm_pais(mensaje)
    if any(k in t for k in _KEYWORDS_INTERCONEXION):
        return True
    if re.search(r"[a-z]\s*[-–>→]\s*[a-z]", t):  # países encadenados: A-B-C, A→B…
        return True
    return False


def analizar_cadena(paises: List[str]) -> Dict[str, Any]:
    """Por cada parámetro: intersección de los rangos de todos los países de la cadena (en
    condiciones de España) y qué país impone la restricción más dura. Marca incompatibilidades."""
    filas: List[Dict[str, Any]] = []
    cuellos: Dict[str, List[str]] = {}
    incompat: List[str] = []
    for pinfo in PARAMETROS_UI:
        slug, label = pinfo["slug"], pinfo["label"]
        unidad_es = _unidad_de_pais(slug, PAIS_BASE) or ""
        lo_max, hi_min = float("-inf"), float("inf")
        pais_lo = pais_hi = None
        hay = False
        for pais in paises:
            r = _rango_en_condiciones_es(slug, pais, unidad_es)
            est = r.get("estado")
            if est in ("sin_datos", "sin_limite", "incomparable"):
                continue  # no impone una restricción numérica comparable
            hay = True
            lo = r["lo"] if r["lo"] is not None else float("-inf")
            hi = r["hi"] if r["hi"] is not None else float("inf")
            if hi < hi_min - 1e-9:
                hi_min, pais_hi = hi, pais
            if lo > lo_max + 1e-9:
                lo_max, pais_lo = lo, pais
        if not hay:
            estado = "sin_datos"
        elif lo_max > hi_min + 1e-9:
            estado = "incompatible"
            incompat.append(label)
        else:
            estado = "ok"
        if estado in ("ok", "incompatible"):
            if pais_hi:
                cuellos.setdefault(pais_hi, []).append(label)
            if pais_lo and pais_lo != pais_hi:
                cuellos.setdefault(pais_lo, []).append(label + " (mín.)")
        filas.append({
            "parametro": slug, "label": label, "unidad": unidad_es,
            "lo": (None if lo_max == float("-inf") else round(lo_max, 4)),
            "hi": (None if hi_min == float("inf") else round(hi_min, 4)),
            "cuello_max": pais_hi, "cuello_min": pais_lo, "estado": estado,
        })
    return {"cadena": paises, "parametros": filas, "cuellos": cuellos, "incompatibilidades": incompat}


def _formatear_cadena(res: Dict[str, Any]) -> str:
    """Respuesta de chat (markdown) del análisis de cadena."""
    paises = res["cadena"]
    L = [f"**Interconexión: {' → '.join(paises)}**", "",
         "Para que un gas circule por toda la cadena debe cumplir el límite de **todos** los "
         "países a la vez. Rango común admisible por parámetro (en condiciones de España) y quién "
         "impone la restricción más estricta (cuello de botella):", "",
         "| Parámetro | Rango común admisible | Cuello de botella |",
         "| --- | --- | --- |"]
    mostrados = 0
    for f in res["parametros"]:
        if f["estado"] not in ("ok", "incompatible"):
            continue
        mostrados += 1
        u = f["unidad"]
        if f["estado"] == "incompatible":
            rango = "🔴 **vacío — incompatible**"
        else:
            a = _coma(f["lo"]) if f["lo"] is not None else "—"
            b = _coma(f["hi"]) if f["hi"] is not None else "—"
            rango = f"{a} / {b} {u}".strip()
        cuello = []
        if f["cuello_max"]:
            cuello.append(f"**{f['cuello_max']}** (máx.)")
        if f["cuello_min"] and f["cuello_min"] != f["cuello_max"]:
            cuello.append(f"**{f['cuello_min']}** (mín.)")
        L.append(f"| {f['label']} | {rango} | {' · '.join(cuello) or '—'} |")
    if not mostrados:
        return (f"**Interconexión: {' → '.join(paises)}**\n\nNinguno de los países de la cadena fija "
                "límites numéricos comparables para los parámetros de calidad, así que no puedo "
                "identificar cuellos de botella.")
    if res["incompatibilidades"]:
        L += ["", "⛔ **Incompatibilidad detectada:** ningún gas puede atravesar toda la cadena, porque "
              "los límites de **" + ", ".join(dict.fromkeys(res["incompatibilidades"])) +
              "** no dejan un rango común entre los países."]
    else:
        L += ["", "✅ **Existe gas admisible** para toda la cadena: el que cumpla los rangos comunes de la tabla."]
    if res["cuellos"]:
        partes = [f"**{pais}** ({', '.join(dict.fromkeys(params))})" for pais, params in res["cuellos"].items()]
        L += ["", "**Cuellos de botella regulatorios** (país con el límite más estricto por parámetro): "
              + "; ".join(partes) + "."]
    L += ["", "_Método: intersección de los límites de cada país normalizados a las condiciones de "
          "España (ISO 13443). El cuello de botella es el país con el límite más estricto en cada parámetro._"]
    return "\n".join(L)


def _interconexion_response(mensaje: str) -> Optional[str]:
    """Si el mensaje pide una interconexión/cadena, devuelve el análisis; si no, None."""
    paises = _extraer_cadena_paises(mensaje)
    if not _es_consulta_interconexion(mensaje, paises):
        return None
    return _formatear_cadena(analizar_cadena(paises))


def _es_consulta_intercambio(texto_norm: str) -> bool:
    """¿Pregunta con qué país se puede intercambiar/compatibilizar el gas?"""
    if "intercambi" in texto_norm or "intercanvi" in texto_norm:
        return True
    pistas_pais = any(p in texto_norm for p in (
        "con que pais", "con qué país", "con que paises", "con qué países",
        "con quien", "con quién", "que pais podria", "qué país podría",
    ))
    return pistas_pais and ("gas" in texto_norm or "compatible" in texto_norm or "intercambi" in texto_norm)


def _es_consulta_restriccion(texto_norm: str) -> bool:
    """¿Pregunta qué país es más restrictivo/amplio que España en un parámetro?"""
    return any(k in texto_norm for k in (
        "restrictiv", "estrict", "exigent", "mas amplio", "más amplio", "mas amplia",
        "más amplia", "permisiv", "laxo", "laxa", "flexible", "menos exigent",
    ))


def _restriccion_vs_espana(parametro: str, texto_norm: str) -> str:
    """#4 — «¿Qué país de la UE tiene un límite de [parámetro] más restrictivo/amplio
    que España?» Clasifica cada país frente a España (en condiciones de España)."""
    display = DISPLAY_MAP.get(parametro, parametro)
    unidad_es = _unidad_de_pais(parametro, PAIS_BASE)
    es = _rango_en_condiciones_es(parametro, PAIS_BASE, unidad_es)
    if es.get("estado") != "ok":
        return f"No tengo un límite numérico de España para {display}, así que no puedo compararlo."
    es_maximo = es.get("lo") is None  # solo cota superior (contaminantes) vs rango
    es_lo = es["lo"] if es["lo"] is not None else float("-inf")
    es_hi = es["hi"] if es["hi"] is not None else float("inf")
    ancho_es = es_hi - es_lo

    quiere_amplio = any(k in texto_norm for k in ("amplio", "amplia", "permisiv", "laxo", "laxa", "flexible", "menos exigent"))
    quiere_restr = any(k in texto_norm for k in ("restrictiv", "estrict", "exigent")) and not quiere_amplio

    def fr(lo, hi):
        return f"{_coma(lo) if lo is not None else '—'} / {_coma(hi) if hi is not None else '—'}"

    paises = [p for p in ALL_COUNTRIES if _norm_pais(p) != _norm_pais(PAIS_BASE)]
    lines = [
        f"**{display}: ¿qué país es más restrictivo/amplio que España?** (base España: {fr(es['lo'], es['hi'])} {unidad_es or ''})",
        "",
        "| País | Límite (en condiciones de España) | Frente a España |",
        "| --- | --- | --- |",
    ]
    mas_restr, mas_amplio = [], []
    for pais in paises:
        r = _rango_en_condiciones_es(parametro, pais, unidad_es)
        est = r.get("estado")
        if est == "sin_datos":
            lines.append(f"| {pais} | — | ⚪ No consta el parámetro |")
            continue
        if est == "incomparable":
            lines.append(f"| {pais} | (no convertible) | ⚪ No comparable |")
            continue
        if est == "sin_limite":
            lines.append(f"| {pais} | Sin límite | 🔼 Más amplio (no lo regula) |")
            mas_amplio.append(pais)
            continue
        lo = r["lo"] if r["lo"] is not None else float("-inf")
        hi = r["hi"] if r["hi"] is not None else float("inf")
        if es_maximo:
            if hi < es_hi - 1e-9:
                verd, lst = "🔽 Más restrictivo (máximo más bajo)", mas_restr
            elif hi > es_hi + 1e-9:
                verd, lst = "🔼 Más amplio (máximo más alto)", mas_amplio
            else:
                verd, lst = "➖ Igual de exigente", None
        else:
            ancho = hi - lo
            if ancho < ancho_es - 1e-9:
                verd, lst = "🔽 Más restrictivo (rango más estrecho)", mas_restr
            elif ancho > ancho_es + 1e-9:
                verd, lst = "🔼 Más amplio (rango más ancho)", mas_amplio
            else:
                verd, lst = "➖ Rango equivalente", None
        if lst is not None:
            lst.append(pais)
        lines.append(f"| {pais} | {fr(r['lo'], r['hi'])} {unidad_es or ''} | {verd} |")

    if quiere_restr:
        foco = f"**Más restrictivos que España:** {', '.join(mas_restr) if mas_restr else 'ninguno'}."
    elif quiere_amplio:
        foco = f"**Más amplios que España:** {', '.join(mas_amplio) if mas_amplio else 'ninguno'}."
    else:
        foco = (f"**Más restrictivos:** {', '.join(mas_restr) or 'ninguno'}. "
                f"**Más amplios:** {', '.join(mas_amplio) or 'ninguno'}.")
    lines += ["", foco]
    return "\n".join(lines)


def _decimales(s: Any) -> int:
    """Nº de decimales de un valor escrito ('10,26' -> 2, '50' -> 0)."""
    t = str(s)
    for sep in (",", "."):
        if sep in t:
            return len(t.split(sep)[-1].strip())
    return 0


def _redondea_coma(v: Optional[float], dec: int) -> str:
    if v is None:
        return "—"
    if dec <= 0:
        return str(int(round(v)))
    return f"{round(v, dec):.{dec}f}".replace(".", ",")


def _parse_notac(n: Any) -> tuple:
    """'(25/0)' -> (25.0, 0.0)  [combustión, volumen]."""
    try:
        a, b = str(n).strip("()").split("/")
        return float(a.replace(",", ".")), float(b.replace(",", "."))
    except Exception:
        return (0.0, 0.0)


def _condiciones_iguales(parametro: str, notac_a: Any, notac_b: Any) -> bool:
    """¿Las condiciones de referencia son equivalentes PARA ESTE parámetro?

    La temperatura del VOLUMEN de referencia solo afecta a magnitudes VOLUMÉTRICAS:
    PCS/Wobbe (MJ/m³) y concentraciones MÁSICAS por volumen (mg/m³ — azufre total,
    H₂S+COS, mercaptanos). El % mol, el ppm, lo adimensional (densidad relativa) y los
    puntos de rocío (°C) NO dependen de la temperatura del volumen. Para PCS/Wobbe
    importa además la temperatura de combustión."""
    ca, va = _parse_notac(notac_a)
    cb, vb = _parse_notac(notac_b)
    es_comb = _slug_param_comb(parametro) in {"pcs", "wobbe"}
    p = _norm_pais(parametro)
    masa_volumen = any(k in p for k in ("s total", "azufre", "h2s", "mercapt", "rsh"))
    if (es_comb or masa_volumen) and va != vb:
        return False
    if es_comb and ca != cb:
        return False
    return True


def _celda_limite(inf: Any, sup: Any, unidad: str, notac: str) -> str:
    i, s = _txt(inf), _txt(sup)
    if _sin_limite(i) and _sin_limite(s):
        return "Sin límite numérico"
    rango = f"{i if not _sin_limite(i) else '—'} - {s if not _sin_limite(s) else '—'}"
    return f"{rango} {unidad}{(' ' + notac) if notac else ''}".strip()


def _limite_convertido_a_es(parametro, pais, inf, sup, u_pa, unidad_es, notac_es, dec):
    """Convierte el límite del país a unidad+condiciones de España (ISO 13443) y lo
    redondea a las cifras de España (Nota 1)."""
    vi, vs = _num_limite(inf), _num_limite(sup)
    if vi is None and vs is None:
        return "Sin límite numérico"

    def conv(v):
        if v is None:
            return None
        if u_pa and unidad_es and _normalize_unit(u_pa) != _normalize_unit(unidad_es):
            c = convertir_unidades(v, u_pa, unidad_es, parametro)
            if "valor_convertido" not in c:
                return "incompat"
            v = c["valor_convertido"]
        if _slug_param_comb(parametro) in {"pcs", "wobbe"} and _norm_pais(pais) != _norm_pais(PAIS_BASE):
            v = convertir_a_condiciones_espana(v, parametro, pais)["valor_convertido"]
        return v

    ci, cs = conv(vi), conv(vs)
    if ci == "incompat" or cs == "incompat":
        return "No convertible de forma determinista"
    a = _redondea_coma(ci, dec) if ci is not None else "—"
    b = _redondea_coma(cs, dec) if cs is not None else "—"
    return f"{a} - {b} {unidad_es} {notac_es}".strip()


def _comparativa_enagas(parametro: str, pais: str) -> str:
    """Módulo 2 (formato Enagás): «¿Cuál es el límite de [parámetro] en [país]
    comparado con España?». Tabla de 5 columnas con el límite del país convertido a
    unidad y condiciones de España (ISO 13443), redondeado a las cifras de España."""
    display = DISPLAY_MAP.get(parametro, parametro)
    es_resp = _consultar_norma(parametro, PAIS_BASE)
    pa_resp = _consultar_norma(parametro, pais)
    if not es_resp.get("matches"):
        return f"No tengo el límite oficial de España para {display}."
    if not pa_resp.get("matches"):
        return f"No tengo el límite oficial de {pais} para {display}."
    es = es_resp["matches"][0]
    unidad_es = es.get("unidad") or ""
    notac_es = es.get("notacion") or "(0/0)"
    es_inf, es_sup = _txt(es.get("limite_inferior")), _txt(es.get("limite_superior"))
    es_celda = _celda_limite(es_inf, es_sup, unidad_es, notac_es)
    ref = es_inf if not _sin_limite(es_inf) else es_sup
    dec = _decimales(ref)

    lines = [
        f"**¿Cuál es el límite de {display} en {pais} comparado con España?**",
        "",
        f"| Parámetro | Límite en España | Límite en {pais} | ¿Unidades y condiciones comparables? | Límite {pais} en condiciones comparables |",
        "| --- | --- | --- | --- | --- |",
    ]
    evidencias = [f"- **España** · {es.get('parametro') or display}: {_cita_oficial(es)}"]
    # Condiciones/matices que el reglamento adjunta a cada límite (p. ej. el O₂ de la
    # EN 16726 es ≤1 % general, pero ≤0,01 %/0,001 % por proceso de evaluación en
    # instalaciones sensibles). Sin esto, la comparación de un solo número engaña.
    notas: list = []
    if es.get("nota"):
        notas.append(f"- **España**: {es['nota']}")
    for m in pa_resp["matches"]:  # Nota 3: una fila por límite (tipos de gas H/L…)
        nombre = str(m.get("parametro") or display).strip()
        u_pa = _txt(m.get("unidad")).strip("()").replace("^3", "³").replace("^2", "²")
        notac_pa = m.get("notacion") or "(0/0)"
        pa_inf, pa_sup = _txt(m.get("limite_inferior")), _txt(m.get("limite_superior"))
        pa_celda = _celda_limite(pa_inf, pa_sup, u_pa, notac_pa)
        misma_unidad = _normalize_unit(u_pa) == _normalize_unit(unidad_es)
        comparable = "Sí" if (misma_unidad and _condiciones_iguales(parametro, notac_pa, notac_es)) else "No"
        conv_celda = _limite_convertido_a_es(parametro, pais, pa_inf, pa_sup, u_pa, unidad_es, notac_es, dec)
        lines.append(f"| {nombre} | {es_celda} | {pa_celda} | {comparable} | {conv_celda} |")
        evidencias.append(f"- **{pais}** · {nombre}: {_cita_oficial(m)}")
        if m.get("nota"):
            notas.append(f"- **{pais}**: {m['nota']}")
    if notas:
        lines += ["", "**Condiciones y matices del reglamento**", ""] + notas
    lines += ["", "**Fuente consultada**", ""] + evidencias
    if _slug_param_comb(parametro) in {"pcs", "wobbe"}:
        lines += ["", "_Conversión de unidad y de condiciones de referencia según ISO 13443:1996; "
                  "valor convertido redondeado a las cifras significativas del límite de España._"]
    return "\n".join(lines)


def _es_consulta_fuente(texto_norm: str) -> bool:
    """¿El usuario pregunta de qué reglamento/norma procede la información?"""
    claves = (
        "reglament", "normativ", "documento", "fuente", "real decreto", "orden ted",
        "directiva", "origen normativo", "base legal", "referencia legal", "que norma",
        "qué norma", "de donde sale", "de dónde sale", "de donde proviene", "de dónde proviene",
        "de donde procede", "de dónde procede", "procede de", "proviene de", "en que se basa",
        "en qué se basa", "de que ley", "de qué ley", "que ley regula", "qué ley regula",
    )
    return any(k in texto_norm for k in claves)


def _pregunta_lista_fuentes(texto_norm: str) -> bool:
    """Pregunta GENERAL por el conjunto de reglamentos (sin un parámetro concreto)."""
    patrones = (
        "que reglamentos", "qué reglamentos", "que normas", "qué normas",
        "que normativas", "qué normativas", "que fuentes", "qué fuentes",
        "que documentos", "qué documentos", "reglamentos usa", "normas usa",
        "fuentes usa", "en que se basa", "en qué se basa", "que reglamento usa",
    )
    return any(p in texto_norm for p in patrones)


def _validate_measurement_gate(session_id: str, mensaje: str) -> Optional[str]:
    texto_norm = mensaje.lower()
    pending = pending_unit_validations.get(session_id)
    if pending and _parse_numeric_value(mensaje) is None:
        unidad_respuesta = _extract_unit_only(mensaje)
        if unidad_respuesta is None:
            return None
        if not _unit_matches_expected(pending["parametro"], unidad_respuesta):
            pending_unit_validations.pop(session_id, None)
            return _incorrect_unit_message(pending["parametro"])
        pending_unit_validations.pop(session_id, None)
        return _evaluar_paises(
            pending["parametro"], pending["valor"], unidad_respuesta,
            pending["paises"], todos=pending.get("todos", False),
        )

    parametro = _normalize_parameter(texto_norm)
    paises = _detectar_paises(texto_norm)        # lista de países pedidos (tolera erratas)
    # "02" aislado significa O₂ (oxígeno), NO el número 2: normalízalo antes de buscar valores.
    mensaje_num = re.sub(r"(?<![\w./,])02(?![\w])", "o2", mensaje)
    valor_con_unidad, unidad_detectada = _extract_numeric_with_unit(mensaje_num)
    valor = valor_con_unidad if valor_con_unidad is not None else _parse_numeric_value(mensaje_num)
    # Si el número y la unidad venían separados ("0.03de % molar"), busca la unidad aparte.
    if unidad_detectada is None:
        unidad_detectada = _extract_unit_only(mensaje)

    # Conversión a CONDICIONES DE REFERENCIA de España (temperatura de combustión),
    # con los factores de la Tabla A.1 de la ISO 13443. Requiere lenguaje explícito de
    # condiciones, así que no interfiere con cumplimiento/límite normales.
    if parametro is not None and _es_consulta_condiciones(texto_norm):
        pais_origen = next((p for p in paises if _norm_pais(p) != _norm_pais(PAIS_BASE)), "Portugal")
        return _responder_condiciones(parametro, valor, unidad_detectada, pais_origen)

    # Compliance: hay parámetro + valor, y se menciona país(es) o hay señal de cumplimiento.
    cue_cumplimiento = any(c in texto_norm for c in [
        "cumple", "paises", "países", "donde", "dónde", "pais", "país", "valido", "válido", "dentro",
    ])
    if parametro is not None and valor is not None and (paises or cue_cumplimiento):
        # FILTRADO ESTRICTO: si hay país(es) explícito(s), solo esos. Si no, todos.
        todos = not paises
        paises_efectivos = paises if paises else list(ALL_COUNTRIES)
        expected = _expected_unit_for_parameter(parametro)
        if expected and unidad_detectada is None:
            pending_unit_validations[session_id] = {
                "parametro": parametro, "paises": paises_efectivos, "todos": todos, "valor": valor,
            }
            return _missing_unit_message(parametro)
        if expected and unidad_detectada is not None and not _unit_matches_expected(parametro, unidad_detectada):
            return _incorrect_unit_message(parametro)
        return _evaluar_paises(parametro, valor, unidad_detectada, paises_efectivos, todos=todos)

    # El usuario plantea un cumplimiento (valor + unidad o señal de "cumple") pero el
    # parámetro no se reconoce → indícalo y ofrece la lista de parámetros disponibles.
    if parametro is None and valor is not None and (unidad_detectada is not None or cue_cumplimiento):
        return _parametro_no_reconocido_message()

    # ¿De qué reglamento/norma procede la información? → fuente documental verificada
    # (ontología). Tiene prioridad sobre comparación/límite (puede mencionar ambos).
    if valor is None and _es_consulta_fuente(texto_norm):
        if parametro is not None:
            paises_f = list(paises) if paises else list(_PAISES_FUENTE)
            return _responder_fuente(parametro, paises_f)
        if paises:  # #1 «¿qué normativa aplican en [país]?» (sin parámetro concreto)
            return _normativa_de_pais(paises)
        if _pregunta_lista_fuentes(texto_norm):
            return _listar_fuentes()

    # #2 «Teniendo en cuenta los límites de España de [parámetro], ¿con qué país
    # podría intercambiar gas?» → intercambiabilidad (solape de rangos vs España).
    if parametro is not None and valor is None and _es_consulta_intercambio(texto_norm):
        return _intercambiabilidad(parametro)

    # #4 «¿Qué país tiene un límite de [parámetro] más restrictivo/amplio que España?»
    if parametro is not None and valor is None and _es_consulta_restriccion(texto_norm):
        return _restriccion_vs_espana(parametro, texto_norm)

    # Comparación de NORMATIVA entre países (sin valor del usuario):
    # "compara el Wobbe entre España y Francia", "diferencia de O2 España vs Portugal"…
    cue_comparar = any(c in texto_norm for c in [
        "compara", "comparar", "comparacion", "comparación", "diferencia",
        "frente a", "versus", " vs ", "enfrenta", "respecto",
    ])
    if parametro is not None and valor is None and (len(paises) >= 2 or (cue_comparar and len(paises) >= 1)):
        otros = [p for p in paises if _norm_pais(p) != _norm_pais(PAIS_BASE)]
        # España vs UN país → formato Enagás (módulo 2): tabla con conversión a condiciones ES.
        if len(otros) == 1:
            return _comparativa_enagas(parametro, otros[0])
        # Varios países → tabla comparativa multi-país.
        paises_efectivos = list(paises) if otros else [PAIS_BASE]
        return _comparar_normativa(parametro, paises_efectivos)

    # Consulta del LÍMITE/valor de un parámetro SIN que el usuario aporte un valor a
    # evaluar → mostrar los límites, NUNCA "cumple/no cumple" (sin un valor no hay nada
    # que cumplir). Cierra el hueco por el que la IA inventaba veredictos de cumplimiento.
    if parametro is not None and valor is None and _es_consulta_limite(texto_norm):
        paises_info = list(paises) if paises else list(ALL_COUNTRIES)
        return _comparar_normativa(parametro, paises_info)

    return None


def _es_consulta_limite(texto_norm: str) -> bool:
    """¿El usuario pregunta por el límite/valor de un parámetro (sin dar un valor a evaluar)?"""
    claves = (
        "limite", "límite", "limites", "límites", "valor", "valores", "rango",
        "maximo", "máximo", "minimo", "mínimo", "cuanto", "cuánto", "umbral",
        "tope", "especificac", "requisito", "requisitos", "que valor", "qué valor",
        "exige", "permite", "permitido", "admite", "admitido", "establece",
        # "¿Cuál es el [parámetro] en [país]?" → mostrar el límite (antes caía en el LLM).
        "cual es", "cuál es", "cuanto vale", "cuánto vale", "cuales son", "cuáles son",
    )
    return any(k in texto_norm for k in claves)


def _is_info_request(text: str) -> bool:
    lowered = text.lower()
    keywords = [
        "limite",
        "límite",
        "requisito",
        "documento",
        "regula",
        "diferencia",
        "todos los límites",
        "todos los limites",
        "monitoriz",
        "parámetro",
        "parametro",
        "origen",
        "cuál es",
        "que es",
        "qué",
    ]
    return any(keyword in lowered for keyword in keywords)


def _format_info_response(
    parametro: str,
    pais: str,
    respuesta: Dict[str, Any],
) -> str:
    matches = respuesta.get("matches", [])
    if not matches:
        return f"No encontré información determinista para '{parametro}' en '{pais}'."

    lines = [
        f"**Consulta sobre {parametro} en {pais}**",
        "",
        "| Parámetro | Límites aplicables | Condiciones de medición |",
        "| --- | --- | --- |",
    ]
    evidencias = []
    for item in matches[:8]:
        parametro_name = item.get("parametro", parametro)
        inferior = item.get("limite_inferior", "-")
        superior = item.get("limite_superior", "-")
        unidad_reg = item.get("unidad") or item.get("unidad_registro") or ""
        unidad_reg = unidad_reg.strip().strip("()")
        condiciones = _normalize_condition_text(item.get("condiciones") or item.get("condiciones de medicion") or item.get("condiciones de medición"))
        if not condiciones:
            condiciones = "No especificadas en el registro"
        if inferior == "-" and superior == "-":
            rango = "Sin límites numéricos definidos"
        else:
            rango = f"{inferior} / {superior}"
            if unidad_reg:
                rango = f"{rango} ({unidad_reg})"
        lines.append(f"| {parametro_name} | {rango} | {condiciones} |")
        evidencias.append(f"- **{parametro_name}**: {_cita_oficial(item)}")
        if item.get("discrepancia"):
            evidencias.append(f"  - ⚠ Discrepancia con el Excel (prevalece la oficial): {item['discrepancia']}")
    if evidencias:
        lines += ["", "**Evidencias (fuente oficial)**", ""] + evidencias
    return "\n".join(lines)


def _fallback_deterministic_response(mensaje: str, session_id: str = "default") -> str:
    texto = mensaje
    texto_norm = texto.lower()

    validation_response = _validate_measurement_gate(session_id, mensaje)
    if validation_response is not None:
        return validation_response

    parametro = _normalize_parameter(texto_norm)
    pais = next((kw for kw in ["espa", "portugal", "francia", "espana", "españa"] if kw in texto_norm), None)
    pais_formateado = _normalize_country(pais) if pais else None

    valor_con_unidad, unidad_detectada = _extract_numeric_with_unit(texto)
    valor = valor_con_unidad if valor_con_unidad is not None else _parse_numeric_value(texto)

    comparison_intent = (
        parametro is not None
        and pais_formateado is not None
        and valor is not None
        and (
            unidad_detectada is not None
            or any(token in texto_norm for token in ("cumple", "válido", "valido", "excede", "dentro", "rango", "compar"))
        )
    )

    if comparison_intent and unidad_detectada and _unit_matches_expected(parametro, unidad_detectada):
        return _evaluar_paises(parametro, valor, unidad_detectada, [pais_formateado])

    if (
        parametro is not None
        and pais_formateado is not None
        and valor is not None
        and not _is_info_request(texto_norm)
    ):
        if unidad_detectada is None:
            return _missing_unit_message(parametro)
        if not _unit_matches_expected(parametro, unidad_detectada):
            return _incorrect_unit_message(parametro)
        # If unit matches but we are here because not comparison_intent? Actually this block runs when not info request.
        # We'll just fall through to default handling (maybe ask for country/param etc.)

    if parametro and pais_formateado and _is_info_request(texto_norm):
        respuesta = _consultar_norma(parametro, pais_formateado)
        if respuesta.get("count", 0) == 0:
            return f"No encontré información específica para '{parametro}' en '{pais_formateado}'."
        return _format_info_response(parametro, pais_formateado, respuesta)

    if parametro and pais_formateado:
        respuesta = _consultar_norma(parametro, pais_formateado)
        if respuesta.get("count", 0) == 0:
            pdf_resultados = buscar_pdfs(query=texto_norm)
            if pdf_resultados["count"] > 0:
                primer_resultado = pdf_resultados["matches"][0]
                return (
                    f"No encontré coincidencia exacta en el Excel/CSV para '{parametro}' en '{pais_formateado}', "
                    f"pero sí encontré información en PDF: {primer_resultado.get('name')} (página {primer_resultado.get('page')}). "
                    f"Extracto: {primer_resultado.get('snippet', '')}"
                )
            return f"No encontré información específica para '{parametro}' en '{pais_formateado}'."
        return _format_info_response(parametro, pais_formateado, respuesta)

    # 1) El usuario pregunta qué puede consultar el chatbot → lista de parámetros.
    if _es_pregunta_capacidades(texto_norm):
        return _mensaje_capacidades()

    # 2) La pregunta NO trata de calidad del gas (aunque mencione un país) → fuera de ámbito.
    if not _es_tema_calidad_gas(texto_norm):
        return _mensaje_fuera_de_ambito()

    # 3) Es de calidad del gas pero no reconocimos el parámetro → indícalo y ofrece la lista.
    return _parametro_no_reconocido_message()


@app.get(
    "/api/status",
    response_model=StatusResponse,
    tags=["Estado"],
    summary="Estado y modo operativo del backend",
    response_description="Modo actual (`ia`/`determinista`) y detalle legible.",
)
@gestionar_errores
async def status_endpoint() -> StatusResponse:
    """Indica si el backend tiene un LLM operativo (`modo: "ia"`) o funciona solo con el
    motor determinista (`modo: "determinista"`).

    El frontend lo usa al cargar para mostrar el indicador de estado. Se calcula una vez al
    arrancar el servidor, según la disponibilidad del proveedor LLM configurado.
    """
    return StatusResponse(modo=backend_mode, detalle=backend_detail)


@app.post(
    "/api/chat",
    response_model=RespuestaChat,
    tags=["Chat"],
    summary="Enviar un mensaje al asistente",
    response_description="Respuesta redactada y el modo con que se resolvió.",
)
@gestionar_errores
@medir_tiempo
async def chat_endpoint(request: PeticionChat) -> RespuestaChat:
    """Procesa un mensaje de chat y devuelve la respuesta del asistente experto.

    **Flujo (a prueba de fallos):**

    1. **Filtro de coherencia** (`_validate_measurement_gate`): valida el mensaje frente al
       historial de la sesión; si detecta una incoherencia, responde en `modo: "determinista"`.
    2. **Sin LLM disponible** → responde con el motor determinista (`modo: "determinista"`).
    3. **Con LLM** → el modelo redacta la respuesta llamando a herramientas deterministas
       para las cifras (`modo: "ia"`). Si el LLM falla (clave inválida, red, límite de uso…),
       **no** se devuelve un error 500: se cae automáticamente al motor determinista.

    Cada turno se registra en el historial de la sesión (`session_id`) para conservar memoria
    entre mensajes.
    """
    validation_response = _validate_measurement_gate(request.session_id, request.mensaje)
    if validation_response is not None:
        _registrar_turno(request.session_id, request.mensaje, validation_response)
        return RespuestaChat(respuesta=validation_response, modo="determinista")

    # Interconexión / cadena: se resuelve SIEMPRE de forma determinista (cero cifras inventadas).
    interconexion = _interconexion_response(request.mensaje)
    if interconexion is not None:
        _registrar_turno(request.session_id, request.mensaje, interconexion)
        return RespuestaChat(respuesta=interconexion, modo="determinista")

    if not provider.is_available():
        respuesta = _fallback_deterministic_response(request.mensaje, request.session_id)
        _registrar_turno(request.session_id, request.mensaje, respuesta)
        return RespuestaChat(respuesta=respuesta, modo="determinista")

    # Si el LLM falla (clave inválida, red, límite…), NO rompemos el chat:
    # caemos al motor determinista en vez de devolver un error 500.
    try:
        texto = responder_con_llm(request.mensaje, request.session_id)
        return RespuestaChat(respuesta=texto, modo="ia")
    except Exception as exc:  # noqa: BLE001
        print(f"[chat] LLM no disponible ({exc}); usando motor determinista.")
        respuesta = _fallback_deterministic_response(request.mensaje, request.session_id)
        _registrar_turno(request.session_id, request.mensaje, respuesta)
        return RespuestaChat(respuesta=respuesta, modo="determinista")


# --- Sección de COMPARATIVA (desplegables) ---------------------------------
class PeticionComparar(BaseModel):
    """Cuerpo de la petición para `POST /api/comparar`."""

    parametro: str = Field(
        ...,
        description=(
            "Parámetro de calidad a comparar. Admite el `slug` técnico (p. ej. `wobbe`, "
            "`pcs`, `s total`, `h2o(rocío)`) o su etiqueta; se normaliza internamente. "
            "Consulta la lista en `GET /api/parametros`."
        ),
        examples=["pcs"],
    )
    paises: List[str] = Field(
        default_factory=list,
        description=(
            "Países a comparar frente a España, que es SIEMPRE la base de referencia y se "
            "añade automáticamente. Lista vacía = solo España. Ver `GET /api/parametros`."
        ),
        examples=[["Francia", "Alemania"]],
    )
    unidad: Optional[str] = Field(
        default=None,
        description=(
            "Unidad de destino a la que convertir los límites (p. ej. `kWh/m³`, `MJ/m³`, "
            "`mg/Nm³`, `ppm`). Si se omite, cada país se muestra en la unidad nativa de su norma."
        ),
        examples=["kWh/m³"],
    )


_EJEMPLO_PARAMETROS = {
    "parametros": [
        {"slug": "wobbe", "label": "Índice de Wobbe", "unidades": ["kWh/m³", "MJ/m³"]},
        {"slug": "pcs", "label": "PCS (Poder Calorífico Superior)", "unidades": ["kWh/m³", "MJ/m³"]},
    ],
    "paises": ["Portugal", "Francia", "Italia", "Alemania", "…"],
    "base": "España",
}

_EJEMPLO_COMPARAR = {
    "parametro": "PCS (Poder Calorífico Superior)",
    "unidad": "kWh/m³",
    "es_combustion": True,
    "filas": [
        {
            "pais": "España", "parametro": "PCS", "limite": "10,26 / 13,26",
            "unidad": "kWh/m³", "condiciones": "comb. 25 °C · med. 0 °C", "es_base": True,
            "limite_espana": None, "factor_iso": None,
            "fuente": "BOE-A-2018-…", "organismo": "MITECO", "fecha": "2018",
            "articulo": "Anexo …", "url": "https://…", "estado": "ok",
            "discrepancia": "", "nota": "",
        },
        {
            "pais": "Francia", "parametro": "PCS", "limite": "10,70 / 12,80",
            "unidad": "kWh/m³", "condiciones": "comb. 0 °C · med. 0 °C", "es_base": False,
            "limite_espana": "10,70 / 12,80", "factor_iso": 1.0,
            "fuente": "GRTgaz …", "organismo": "GRTgaz", "fecha": "2023",
            "articulo": "…", "url": "https://…", "estado": "ok",
            "discrepancia": "", "nota": "",
        },
    ],
    "notas_iso": [],
}

_EJEMPLO_MATRIZ = {
    "paises": ["España", "Portugal", "Francia", "…"],
    "parametros": [{"slug": "wobbe", "label": "Índice de Wobbe", "unidad": "kWh/m³"}],
    "filas": [
        {
            "pais": "España",
            "celdas": {"wobbe": {"valor": "13,40 / 16,06", "nivel": "base", "flag": False, "flag_desc": ""}},
        },
        {
            "pais": "Francia",
            "celdas": {"wobbe": {"valor": "13,64 / 15,70", "nivel": "restrictivo", "flag": True,
                                  "flag_desc": "condiciones (0/0) → convertido a condiciones de España"}},
        },
    ],
    "unidad_nota": "Valores normalizados a unidad y condiciones de España (ISO 13443 para PCS/Wobbe).",
}


@app.get(
    "/api/parametros",
    tags=["Comparativa"],
    summary="Catálogo de parámetros y países",
    response_description="Parámetros (con unidades disponibles), países comparables y país base.",
    responses={200: {"content": {"application/json": {"example": _EJEMPLO_PARAMETROS}}}},
)
@gestionar_errores
async def parametros_endpoint() -> Dict[str, Any]:
    """Devuelve el catálogo para poblar los desplegables del frontend.

    - **`parametros`**: lista de objetos `{slug, label, unidades[]}` — los 10 parámetros de
      calidad y las unidades a las que se puede convertir cada uno.
    - **`paises`**: países/regiones comparables (España se excluye porque es la base).
    - **`base`**: país de referencia (`"España"`), frente al que se normaliza todo.
    """
    return {"parametros": PARAMETROS_UI, "paises": PAISES_UI, "base": PAIS_BASE}


@app.post(
    "/api/comparar",
    tags=["Comparativa"],
    summary="Comparación puntual de un parámetro entre países",
    response_description="Filas comparativas (una por país/norma) y notas de normalización ISO.",
    responses={200: {"content": {"application/json": {"example": _EJEMPLO_COMPARAR}}}},
)
@gestionar_errores
async def comparar_endpoint(req: PeticionComparar) -> Dict[str, Any]:
    """Compara un parámetro de calidad entre España (base) y los países indicados.

    Cada elemento de **`filas`** describe el límite de un país en su norma:

    - **`limite`** / **`unidad`** / **`condiciones`**: rango `inferior / superior` en la unidad
      pedida (o la nativa) y las condiciones de referencia (temperaturas de combustión y medida).
    - **`limite_espana`**: el mismo rango **normalizado a las condiciones de España**
      (ISO 13443, Tabla A.1, para PCS/Wobbe; corrección de volumen para concentraciones másicas).
      Es `null` para el país base.
    - **`factor_iso`**: factor de conversión aplicado al normalizar (solo PCS/Wobbe, si ≠ 1).
    - **Trazabilidad**: `fuente`, `organismo`, `fecha`, `articulo`, `url`, `estado`, `nota`,
      `discrepancia` — de dónde procede el dato.

    **`notas_iso`** recoge las conversiones realizadas (mensajes legibles).
    """
    slug = _normalize_parameter((req.parametro or "").lower()) or (req.parametro or "").lower()
    return comparar_estructurado(slug, req.paises, req.unidad or "")


@app.get(
    "/api/matriz",
    tags=["Comparativa"],
    summary="Matriz comparativa (heatmap) de todos los países × parámetros",
    response_description="Rejilla países × parámetros con nivel (color) y flag metodológico por celda.",
    responses={200: {"content": {"application/json": {"example": _EJEMPLO_MATRIZ}}}},
)
@gestionar_errores
async def matriz_endpoint() -> Dict[str, Any]:
    """Matriz comparativa completa (heatmap): filas = países, columnas = los 10 parámetros,
    todo normalizado a la unidad y condiciones de España.

    - **`parametros`**: cabeceras `{slug, label, unidad}` (la unidad es la de España).
    - **`filas`**: una por país, con `celdas[slug]` = `{valor, nivel, flag, flag_desc}`:
        - **`valor`**: rango normalizado `inferior / superior` (o `—`, `Sin límite`, `≠ unidades`).
        - **`nivel`**: posición frente a España, para el color — `base`, `restrictivo`, `amplio`,
          `igual`, `sin_ref`, `sin_dato`, `sin_limite`, `incomparable`.
        - **`flag`** / **`flag_desc`**: `true` si hubo diferencia metodológica (unidad o condiciones
          distintas) que obligó a convertir; el texto explica qué se convirtió.
    - **`unidad_nota`**: aclaración sobre la normalización aplicada.

    No recibe parámetros: recalcula la matriz completa a partir de la ontología.
    """
    return matriz_comparativa()


# ============================================================================
# ANÁLISIS DE UN GAS CONCRETO (validación de calidad país a país)
# ---------------------------------------------------------------------------
# El usuario introduce la composición/medidas de un gas y el sistema responde,
# para cada jurisdicción, si CUMPLE / está en ZONA DE ALERTA / NO CUMPLE / no
# tiene límite. Reutiliza el mismo motor que la comparativa/heatmap: se lleva
# tanto el límite de cada país como el valor del usuario a la UNIDAD y
# CONDICIONES de España (ISO 13443) y se comparan de forma determinista.
# ============================================================================

MARGEN_ALERTA = 0.10        # 10 % de la banda/límite → "zona de alerta" (proximidad)
MARGEN_ALERTA_TEMP = 2.0    # margen absoluto para rocíos (°C/K/°F): el % no aplica a temperaturas
_LABEL_PARAM = {p["slug"]: p["label"] for p in PARAMETROS_UI}
_ROCIO_SLUGS = {"h2o(rocío)", "hc(rocío)"}
_SEVERIDAD = {"cumple": 1, "alerta": 2, "no_cumple": 3}


class ComponenteGas(BaseModel):
    """Un componente o medida del gas a validar."""

    parametro: str = Field(
        ...,
        description=(
            "Parámetro (slug o etiqueta): `co2`, `o2`, `h2s+cos`, `s total`, `rsh`, `pcs`, "
            "`wobbe`, `densidad relativa`, `h2o(rocío)`, `hc(rocío)`. Los componentes NO "
            "normativos (`ch4`, `n2`…) se aceptan pero se marcan como informativos y no se validan."
        ),
        examples=["co2"],
    )
    valor: float = Field(..., description="Valor medido del componente.", examples=[2.6])
    unidad: Optional[str] = Field(
        None,
        description=(
            "Unidad del valor (`% molar`, `ppm`, `mg/Nm³`, `kWh/m³`, `°C`…). Si se omite, se "
            "asume la unidad de referencia de España para ese parámetro."
        ),
        examples=["% molar"],
    )


class PeticionAnalisisGas(BaseModel):
    """Cuerpo de la petición para `POST /api/analizar-gas`."""

    componentes: List[ComponenteGas] = Field(
        ..., description="Componentes/medidas del gas a validar."
    )
    paises: Optional[List[str]] = Field(
        None,
        description="Países a evaluar. Si se omite, se evalúan las 21 jurisdicciones (España incluida).",
        examples=[["España", "Francia", "Alemania"]],
    )
    base_pcs: Optional[str] = Field(
        "España",
        description=(
            "Condiciones de referencia asumidas para el PCS/Wobbe introducidos por el usuario. "
            "Por defecto España (combustión 0 °C). Solo afecta a PCS/Wobbe."
        ),
        examples=["España"],
    )
    margen_alerta: Optional[float] = Field(
        None,
        description="Fracción (0–1) para la zona de alerta por proximidad al límite. Por defecto 0.10 (10 %).",
        examples=[0.10],
    )
    tipo_gas: Optional[str] = Field(
        "gas_natural",
        description=(
            "Tipo de gas a validar: `gas_natural` (por defecto, 21 jurisdicciones) o "
            "`biometano` (especificación UE EN 16723-1, jurisdicción única)."
        ),
        examples=["gas_natural", "biometano"],
    )


def _valor_a_condiciones_es(
    valor: Any, unidad_user: str, unidad_es: str, slug: str, base_pcs: str
) -> Optional[float]:
    """Lleva el valor del usuario a la UNIDAD y CONDICIONES de España, para poder compararlo
    con el límite de cada país (que también se normaliza a España). Devuelve None si las
    unidades no son convertibles de forma determinista (→ no evaluable)."""
    if valor is None:
        return None
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return None
    u_src = (unidad_user or unidad_es or "").strip()
    if u_src and unidad_es and _normalize_unit(u_src) != _normalize_unit(unidad_es):
        conv = convertir_unidades(v, u_src, unidad_es, slug)
        if "valor_convertido" not in conv:
            return None  # unidades incompatibles
        v = conv["valor_convertido"]
    # Condiciones de referencia (ISO 13443): solo PCS/Wobbe dependen de la combustión.
    if _slug_param_comb(slug) in {"pcs", "wobbe"} and base_pcs and _norm_pais(base_pcs) != _norm_pais(PAIS_BASE):
        cr = convertir_a_condiciones_espana(v, slug, base_pcs)
        v = cr.get("valor_convertido", v)
    return v


def _estado_desde_rango(v: Optional[float], rango: Dict[str, Any], slug: str, margen: float = MARGEN_ALERTA) -> str:
    """Decide el estado de un parámetro comparando el valor del usuario `v` con el rango
    normalizado a España `{estado, lo, hi}`. Estados: cumple · alerta · no_cumple ·
    sin_limite · no_evaluable · sin_datos."""
    est = rango.get("estado")
    if est in ("sin_datos", "sin_dato"):
        return "sin_datos"
    if est == "incomparable":
        return "no_evaluable"
    if est == "sin_limite":
        return "sin_limite"
    # est == "ok"
    if v is None:
        return "no_evaluable"
    lo, hi = rango.get("lo"), rango.get("hi")
    if lo is None and hi is None:
        return "sin_limite"
    if hi is not None and v > hi:
        return "no_cumple"
    if lo is not None and v < lo:
        return "no_cumple"
    # Dentro de los límites → ¿zona de alerta por proximidad?
    if slug in _ROCIO_SLUGS:  # escala de temperatura: margen absoluto, no porcentual
        m = MARGEN_ALERTA_TEMP
        if (hi is not None and abs(hi - v) < m) or (lo is not None and abs(v - lo) < m):
            return "alerta"
        return "cumple"
    if lo is not None and hi is not None:
        banda = hi - lo
        if banda > 0 and ((hi - v) < margen * banda or (v - lo) < margen * banda):
            return "alerta"
        return "cumple"
    if hi is not None:  # solo máximo
        return "alerta" if (hi > 0 and v >= hi * (1 - margen)) else "cumple"
    # solo mínimo
    return "alerta" if (lo > 0 and v <= lo * (1 + margen)) else "cumple"


def _detalle_estado(estado: str, v: Optional[float], rango: Dict[str, Any], unidad: str) -> str:
    """Explicación legible del estado de un parámetro."""
    lo, hi = rango.get("lo"), rango.get("hi")
    u = f" {unidad}" if unidad else ""
    if estado == "no_cumple":
        if hi is not None and v is not None and v > hi:
            return f"supera el máximo ({_coma(hi)}{u})"
        if lo is not None and v is not None and v < lo:
            return f"no alcanza el mínimo ({_coma(lo)}{u})"
        return "fuera de los límites"
    return {
        "alerta": "cumple, pero cerca del límite",
        "cumple": "dentro de los límites",
        "sin_limite": "este país no fija un límite numérico para este parámetro",
        "no_evaluable": "no comparable de forma determinista (unidades/condiciones)",
        "sin_datos": "sin dato en la fuente para este país",
    }.get(estado, "")


def _fmt_limite_pais(m: Optional[Dict[str, Any]]) -> str:
    """Límite nativo de un país para mostrarlo (rango, ≤ máx, ≥ mín o 'sin límite')."""
    if not m:
        return "sin dato"
    ni = _num_limite(_txt(m.get("limite_inferior")))
    ns = _num_limite(_txt(m.get("limite_superior")))
    if ni is None and ns is None:
        return "sin límite"
    if ni is not None and ns is not None:
        return f"{_coma(ni)} / {_coma(ns)}"
    if ns is not None:
        return f"≤ {_coma(ns)}"
    return f"≥ {_coma(ni)}"


def _veredicto_pais(parametros: List[Dict[str, Any]]) -> str:
    """Veredicto del país = peor severidad entre sus parámetros evaluables."""
    peor = max((_SEVERIDAD.get(p["estado"], 0) for p in parametros), default=0)
    if peor == 0:
        return "sin_datos"
    return {1: "cumple", 2: "alerta", 3: "no_cumple"}[peor]


def _resumen_analisis(resultados: List[Dict[str, Any]]) -> str:
    """Frase-resumen: agrupa países por veredicto, con el parámetro que lo motiva."""
    grupos: Dict[str, List[str]] = {"no_cumple": [], "alerta": [], "cumple": []}
    for r in resultados:
        v = r["veredicto"]
        if v not in grupos:
            continue
        motivos = [p["label"] for p in r["parametros"] if p["estado"] == v] if v != "cumple" else []
        grupos[v].append(r["pais"] + (f" ({', '.join(dict.fromkeys(motivos))})" if motivos else ""))
    partes = []
    if grupos["cumple"]:
        partes.append("Cumple: " + ", ".join(grupos["cumple"]))
    if grupos["alerta"]:
        partes.append("Alerta: " + ", ".join(grupos["alerta"]))
    if grupos["no_cumple"]:
        partes.append("No cumple: " + ", ".join(grupos["no_cumple"]))
    return " · ".join(partes) if partes else "Sin datos evaluables."


def analizar_gas(
    componentes: List[Any], paises: Optional[List[str]] = None,
    base_pcs: str = "España", margen: float = MARGEN_ALERTA,
    tipo_gas: str = "gas_natural",
) -> Dict[str, Any]:
    """Valida un gas concreto contra la normativa de cada país. Devuelve, por país, un
    veredicto (cumple/alerta/no_cumple/sin_datos) y el detalle por parámetro con su cita.

    `tipo_gas` por defecto "gas_natural" → comportamiento idéntico al actual (21 países).
    "biometano" (EN 16723-1 + red FR) e "hidrogeno" (ISO 14687 Grade D) usan su propio
    catálogo y jurisdicciones; al no ser matriz por país NO se normalizan condiciones."""
    es_gn = tipo_gas == "gas_natural"
    normalizadores = {"biometano": _normalize_parameter_biometano,
                      "hidrogeno": _normalize_parameter_hidrogeno}
    normalizar = normalizadores.get(tipo_gas, _normalize_parameter)
    labels = {p["slug"]: p["label"] for p in CATALOGO_POR_GAS.get(tipo_gas, PARAMETROS_UI)}
    orden_paises = paises or list(JURISDICCIONES_POR_GAS.get(tipo_gas, PAISES_MATRIZ))
    # Base de referencia para la unidad: España (gas natural) o la 1.ª jurisdicción del gas.
    base_ref = PAIS_BASE if es_gn else orden_paises[0]
    comp_out: List[Dict[str, Any]] = []
    evaluables: List[Dict[str, Any]] = []
    for c in componentes:
        parametro_in = str(getattr(c, "parametro", "") or "").strip()
        valor = getattr(c, "valor", None)
        unidad_user = str(getattr(c, "unidad", "") or "").strip()
        slug = normalizar(parametro_in.lower())
        if not slug:  # componente no normativo (CH₄, N₂…): informativo, no se valida
            comp_out.append({"parametro": parametro_in, "label": parametro_in, "valor": valor,
                             "unidad": unidad_user, "informativo": True})
            continue
        label = labels.get(slug, _LABEL_PARAM.get(slug, DISPLAY_MAP.get(slug, slug)))
        es = fuente_oficial.consultar(slug, base_ref, tipo_gas)
        unidad_es = ((es["matches"][0].get("unidad") if es.get("matches") else "") or unidad_user or "")
        # Gas natural: normaliza el valor del usuario a las condiciones de España (ISO 13443).
        # Biometano/hidrógeno: una sola especificación → no hay normalización entre jurisdicciones.
        v_norm = _valor_a_condiciones_es(valor, unidad_user, unidad_es, slug, base_pcs) if es_gn else valor
        comp_out.append({"parametro": slug, "label": label, "valor": valor,
                         "unidad": unidad_user or unidad_es, "informativo": False})
        evaluables.append({"slug": slug, "label": label, "valor": valor,
                           "unidad_user": unidad_user or unidad_es, "unidad_es": unidad_es, "v_norm": v_norm})

    resultados: List[Dict[str, Any]] = []
    for pais in orden_paises:
        params_res: List[Dict[str, Any]] = []
        for e in evaluables:
            resp = _consultar_norma(e["slug"], pais, tipo_gas)
            m = resp["matches"][0] if resp.get("matches") else None
            rango = _rango_de_match(m, e["slug"], pais, e["unidad_es"]) if m else {"estado": "sin_datos"}
            estado = _estado_desde_rango(e["v_norm"], rango, e["slug"], margen)
            params_res.append({
                "parametro": e["slug"], "label": e["label"],
                "valor_usuario": e["valor"], "unidad_usuario": e["unidad_user"],
                "valor_evaluado": (round(e["v_norm"], 4) if isinstance(e["v_norm"], float) else e["v_norm"]),
                "unidad_evaluada": e["unidad_es"],
                "limite": _fmt_limite_pais(m), "unidad_limite": (_txt(m.get("unidad")) if m else ""),
                "condiciones": ((_txt(m.get("condiciones")) or _txt(m.get("notacion"))) if m else ""),
                "estado": estado, "detalle": _detalle_estado(estado, e["v_norm"], rango, e["unidad_es"]),
                "fuente": (m.get("documento") if m else "") or "",
                "organismo": (m.get("organismo") if m else "") or "",
                "articulo": (m.get("articulo") if m else "") or "",
                "url": ((m.get("url") or m.get("pdf")) if m else "") or "",
                "estado_fuente": (m.get("estado") if m else "") or "",
                "nota": (m.get("nota") if m else "") or "",
            })
        resultados.append({"pais": JURISDICCION_DISPLAY.get(pais, pais),
                           "veredicto": _veredicto_pais(params_res), "parametros": params_res})

    return {
        "componentes": comp_out, "paises": resultados,
        "resumen": _resumen_analisis(resultados),
        "margen_alerta": margen, "base_pcs": base_pcs, "tipo_gas": tipo_gas,
    }


_EJEMPLO_ANALISIS = {
    "componentes": [
        {"parametro": "co2", "label": "CO₂", "valor": 2.6, "unidad": "% molar", "informativo": False},
        {"parametro": "ch4", "label": "ch4", "valor": 97, "unidad": "% molar", "informativo": True},
    ],
    "paises": [
        {"pais": "España", "veredicto": "no_cumple", "parametros": [
            {"parametro": "co2", "label": "CO₂", "valor_usuario": 2.6, "unidad_usuario": "% molar",
             "valor_evaluado": 2.6, "unidad_evaluada": "% mol", "limite": "≤ 2,5", "unidad_limite": "% mol",
             "condiciones": "(0/0)", "estado": "no_cumple", "detalle": "supera el máximo (2,5 % mol)",
             "fuente": "ORDEN_TED_181_2025", "organismo": "MITECO", "articulo": "Anexo…", "url": "https://…",
             "estado_fuente": "VERIFICADO", "nota": ""}]},
    ],
    "resumen": "No cumple: España (CO₂), Alemania (CO₂)…",
    "margen_alerta": 0.1, "base_pcs": "España",
}


@app.post(
    "/api/analizar-gas",
    tags=["Comparativa"],
    summary="Validar un gas concreto contra la normativa de cada país",
    response_description="Veredicto por país (cumple/alerta/no cumple) con el detalle por parámetro y su cita.",
    responses={200: {"content": {"application/json": {"example": _EJEMPLO_ANALISIS}}}},
)
@gestionar_errores
@medir_tiempo
async def analizar_gas_endpoint(req: PeticionAnalisisGas) -> Dict[str, Any]:
    """Valida la composición/medidas de un gas concreto contra los límites regulatorios de
    cada jurisdicción.

    - **`componentes`**: cada `{parametro, valor, unidad}` se evalúa contra el límite de cada
      país. Los componentes no normativos (CH₄, N₂…) se devuelven como `informativo: true` y
      **no** se validan (el sistema no inventa PCS/Wobbe a partir de la composición).
    - **Veredicto por país** (`cumple` · `alerta` · `no_cumple` · `sin_datos`): la peor
      severidad entre sus parámetros. `alerta` = cumple pero dentro del margen de proximidad
      al límite (por defecto 10 %; configurable con `margen_alerta`).
    - Todo se normaliza a la unidad y condiciones de España (ISO 13443 para PCS/Wobbe) antes
      de comparar, igual que la comparativa y el heatmap. Cada parámetro incluye su cita oficial.
    """
    return analizar_gas(
        req.componentes,
        req.paises,
        req.base_pcs or "España",
        req.margen_alerta if req.margen_alerta is not None else MARGEN_ALERTA,
        req.tipo_gas or "gas_natural",
    )


# ---------------------------------------------------------------------------
# Comparativa de límites por gas (secciones "Biometano" e "Hidrógeno")
# ---------------------------------------------------------------------------
def tabla_limites_gas(tipo_gas: str) -> Dict[str, Any]:
    """Tabla de límites de la especificación de biometano o hidrógeno: parámetros ×
    jurisdicciones, cada celda con su límite y cita. Es la sección "Comparativa" dedicada
    a cada gas. NO toca la comparativa/matriz del gas natural (reutiliza sus helpers)."""
    catalogo = CATALOGO_POR_GAS.get(tipo_gas, [])
    codigos = list(JURISDICCIONES_POR_GAS.get(tipo_gas, []))
    jurisdicciones = [{"codigo": c, "nombre": JURISDICCION_DISPLAY.get(c, c)} for c in codigos]
    parametros: List[Dict[str, Any]] = []
    for p in catalogo:
        celdas: List[Dict[str, Any]] = []
        for c in codigos:
            resp = _consultar_norma(p["slug"], c, tipo_gas)
            m = resp["matches"][0] if resp.get("matches") else None
            celdas.append({
                "jurisdiccion": c,
                "limite": _fmt_limite_pais(m) if m else "—",
                "unidad": (_txt(m.get("unidad")) if m else ""),
                "estado": (m.get("estado") if m else "") or "",
                "fuente": (m.get("documento") if m else "") or "",
                "articulo": (m.get("articulo") if m else "") or "",
                "url": ((m.get("url") or m.get("pdf")) if m else "") or "",
                "expresion": (m.get("expresion_original") if m else "") or "",
                "nota": (m.get("nota") if m else "") or "",
            })
        parametros.append({"slug": p["slug"], "label": p["label"], "celdas": celdas})
    return {"tipo_gas": tipo_gas, "jurisdicciones": jurisdicciones, "parametros": parametros}


@app.get(
    "/api/tabla-limites",
    tags=["Comparativa"],
    summary="Tabla de límites de una especificación de gas (biometano o hidrógeno)",
    response_description="Parámetros × jurisdicciones con el límite y la cita de cada celda.",
)
@gestionar_errores
@medir_tiempo
async def tabla_limites_endpoint(tipo_gas: str = "biometano") -> Dict[str, Any]:
    """Devuelve la tabla de límites de la spec de **biometano** (EN 16723-1 + red GRTgaz)
    o **hidrógeno** (ISO 14687 Grade D). Alimenta las secciones dedicadas de la web."""
    tg = tipo_gas if tipo_gas in ("biometano", "hidrogeno") else "biometano"
    return tabla_limites_gas(tg)


# ============================================================================
# EXPORTACIÓN DE INFORMES COMPARATIVOS (Excel / PDF)
# ---------------------------------------------------------------------------
# Genera un informe con la matriz comparativa (países × parámetros) de las
# jurisdicciones seleccionadas, en Excel (.xlsx, openpyxl) o PDF (xhtml2pdf).
# Serializa los MISMOS datos que la matriz de la web (cero cifras nuevas).
# ============================================================================

# Colores de nivel del heatmap (mismos que index.html), sin '#', para openpyxl/PDF.
_NIVEL_FILL = {
    "base": "FFFFFF", "igual": "E7F4EC", "restrictivo": "E2EDFB", "amplio": "FBEEDD",
    "sin_limite": "F1F3F5", "sin_dato": "F8F9FA", "incomparable": "FBE7E6", "sin_ref": "F8F9FA",
}
_NIVEL_ETIQUETA = {
    "base": "España (base)", "igual": "Igual de exigente", "restrictivo": "Más restrictivo",
    "amplio": "Más amplio", "sin_limite": "Sin límite", "sin_dato": "Sin dato",
    "incomparable": "No comparable", "sin_ref": "Sin referencia",
}


class PeticionExportar(BaseModel):
    """Cuerpo de la petición para `POST /api/exportar-matriz`."""

    paises: Optional[List[str]] = Field(
        None,
        description="Jurisdicciones a incluir en el informe. España se incluye siempre (base). "
                    "Si se omite, se exportan las 21.",
        examples=[["Francia", "Alemania", "Italia"]],
    )
    formato: str = Field(
        "xlsx", description="Formato del informe: `xlsx` (Excel) o `pdf`.", examples=["xlsx"]
    )


def _esc_html(v: Any) -> str:
    """Escape mínimo para insertar texto en el HTML del PDF."""
    return (_txt(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _matriz_para_exportar(paises: Optional[List[str]] = None) -> Dict[str, Any]:
    """Matriz comparativa filtrada a las jurisdicciones pedidas (España siempre incluida)."""
    data = matriz_comparativa()
    if paises:
        seleccion = {_norm_pais(p) for p in paises}
        seleccion.add(_norm_pais(PAIS_BASE))  # España es la base: siempre presente
        data = dict(data)
        data["filas"] = [f for f in data["filas"] if _norm_pais(f["pais"]) in seleccion]
        data["paises"] = [p for p in data["paises"] if _norm_pais(p) in seleccion]
    return data


def _filas_ordenadas_export(filas: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """España primero (base); el resto, alfabético."""
    return sorted(
        filas,
        key=lambda f: (0 if _norm_pais(f["pais"]) == _norm_pais(PAIS_BASE) else 1, f["pais"]),
    )


def _matriz_a_xlsx(data: Dict[str, Any]) -> bytes:
    """Serializa la matriz a un Excel (.xlsx) con celdas coloreadas por nivel."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cols = data.get("parametros", [])
    filas = _filas_ordenadas_export(data.get("filas", []))
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativa"

    ws["A1"] = "Comparativa de calidad de gas natural"
    ws["A1"].font = Font(bold=True, size=14, color="013A57")
    ws["A2"] = data.get("unidad_nota", "")
    ws["A2"].font = Font(italic=True, size=9, color="5D7082")

    hdr = 4
    navy = PatternFill("solid", fgColor="013A57")
    blanco = Font(bold=True, color="FFFFFF")
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(hdr, 1, "País \\ Parámetro")
    for j, c in enumerate(cols, start=2):
        etq = c.get("label", "") + (f" ({c['unidad']})" if c.get("unidad") else "")
        ws.cell(hdr, j, etq)
    for j in range(1, len(cols) + 2):
        cel = ws.cell(hdr, j)
        cel.fill = navy
        cel.font = blanco
        cel.alignment = centro

    fila = hdr + 1
    for f in filas:
        ws.cell(fila, 1, f["pais"]).font = Font(bold=True)
        for j, c in enumerate(cols, start=2):
            celda = f.get("celdas", {}).get(c["slug"], {})
            valor = celda.get("valor", "—")
            if celda.get("flag"):
                valor = "⚠ " + valor
            out = ws.cell(fila, j, valor)
            out.fill = PatternFill("solid", fgColor=_NIVEL_FILL.get(celda.get("nivel"), "FFFFFF"))
            out.alignment = centro
        fila += 1

    ws.cell(fila + 1, 1, "Nota: ⚠ = unidad o condiciones distintas a España (convertido). "
                        "Valores normalizados a la unidad y condiciones de España (ISO 13443 para PCS/Wobbe).")
    ws.cell(fila + 1, 1).font = Font(italic=True, size=8, color="5D7082")

    ws.column_dimensions["A"].width = 22
    for j in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 17
    ws.freeze_panes = "B5"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _matriz_a_pdf(data: Dict[str, Any]) -> bytes:
    """Serializa la matriz a un PDF (A4 apaisado) con celdas coloreadas por nivel."""
    from io import BytesIO
    from xhtml2pdf import pisa

    cols = data.get("parametros", [])
    filas = _filas_ordenadas_export(data.get("filas", []))
    th = "".join(
        f"<th>{_esc_html(c.get('label',''))}"
        f"{('<br/>' + _esc_html(c['unidad'])) if c.get('unidad') else ''}</th>"
        for c in cols
    )
    cuerpo = ""
    for f in filas:
        celdas = ""
        for c in cols:
            celda = f.get("celdas", {}).get(c["slug"], {})
            valor = celda.get("valor", "—")
            if celda.get("flag"):
                valor = "&#9888; " + valor
            bg = _NIVEL_FILL.get(celda.get("nivel"), "FFFFFF")
            celdas += f'<td bgcolor="#{bg}">{_esc_html(valor)}</td>'
        cuerpo += f'<tr><td class="pais">{_esc_html(f["pais"])}</td>{celdas}</tr>'

    leyenda = " · ".join(
        f'<font bgcolor="#{_NIVEL_FILL[k]}">&nbsp;{_esc_html(v)}&nbsp;</font>'
        for k, v in [("base", "España (base)"), ("igual", "Igual"), ("restrictivo", "Más restrictivo"),
                     ("amplio", "Más amplio"), ("sin_limite", "Sin límite"), ("incomparable", "No comparable"),
                     ("sin_dato", "Sin dato")]
    )
    html = f"""<html><head><meta charset="utf-8"/><style>
      @page {{ size: A4 landscape; margin: 1.2cm; }}
      body {{ font-family: Helvetica; font-size: 8pt; color:#1b2a38; }}
      h1 {{ color:#013a57; font-size:14pt; margin:0 0 3px; }}
      p.sub {{ color:#5d7082; font-size:8pt; margin:0 0 10px; }}
      table {{ border-collapse: collapse; width:100%; }}
      th {{ background:#013a57; color:#ffffff; font-size:7pt; padding:4px; border:1px solid #dde4ea; }}
      td {{ font-size:7pt; padding:4px; border:1px solid #dde4ea; text-align:center; }}
      td.pais {{ text-align:left; font-weight:bold; background:#f5f8fa; }}
      p.leg {{ font-size:7pt; color:#5d7082; margin-top:10px; }}
    </style></head><body>
      <h1>Comparativa de calidad de gas natural</h1>
      <p class="sub">{_esc_html(data.get('unidad_nota',''))}</p>
      <table><thead><tr><th>País \\ Parámetro</th>{th}</tr></thead><tbody>{cuerpo}</tbody></table>
      <p class="leg">{leyenda} · &#9888; unidad/condiciones distintas (convertido a España).</p>
    </body></html>"""

    buf = BytesIO()
    pisa.CreatePDF(src=html, dest=buf, encoding="utf-8")
    return buf.getvalue()


@app.post(
    "/api/exportar-matriz",
    tags=["Comparativa"],
    summary="Exportar la matriz comparativa a Excel o PDF",
    response_description="Fichero binario (`.xlsx` o `.pdf`) con la comparativa de las jurisdicciones seleccionadas.",
    responses={200: {"content": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {},
        "application/pdf": {},
    }}},
)
@gestionar_errores
@medir_tiempo
async def exportar_matriz_endpoint(req: PeticionExportar) -> Response:
    """Genera un informe descargable con la matriz comparativa (países × parámetros) de las
    jurisdicciones seleccionadas. `formato`: `xlsx` (Excel) o `pdf`. España se incluye siempre
    como base. No genera cifras nuevas: serializa los mismos datos que la matriz de la web."""
    data = _matriz_para_exportar(req.paises)
    if (req.formato or "").lower() == "pdf":
        contenido = _matriz_a_pdf(data)
        media, nombre = "application/pdf", "comparativa_gas.pdf"
    else:
        contenido = _matriz_a_xlsx(data)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        nombre = "comparativa_gas.xlsx"
    return Response(
        content=contenido, media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
